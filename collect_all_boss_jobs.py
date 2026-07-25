import argparse
import asyncio
import json
import random
import re
import shutil
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from playwright.async_api import async_playwright


CDP_URL = "http://127.0.0.1:9222"
JOB_URL_MARKER = "/weijd/weijd-job/"

URL_PATTERN = re.compile(r"https?://[^\s<>\"']+")
JOB_ID_PATTERN = re.compile(r"/weijd-job/([^?/#]+)")


def clean_inline(text: str | None) -> str:
    if not text:
        return ""

    return re.sub(r"\s+", " ", text.replace("\xa0", " ")).strip()


def clean_multiline(text: str | None) -> str:
    if not text:
        return ""

    lines = []

    for line in text.replace("\xa0", " ").splitlines():
        line = re.sub(r"[ \t]+", " ", line).strip()

        if line:
            lines.append(line)

    return "\n".join(lines)


def split_middle_dot(text: str) -> list[str]:
    if not text:
        return []

    return [
        part.strip()
        for part in re.split(r"\s*[·•]\s*", text)
        if part.strip()
    ]


def get_job_id(url: str) -> str:
    match = JOB_ID_PATTERN.search(url)

    if not match:
        raise ValueError(f"无法从链接中提取岗位ID：{url}")

    return match.group(1)


def read_job_urls(excel_path: Path) -> list[dict]:
    if not excel_path.exists():
        raise FileNotFoundError(
            f"找不到Excel文件：{excel_path.resolve()}"
        )

    workbook = load_workbook(
        excel_path,
        read_only=True,
        data_only=True,
    )

    records = []
    seen_job_ids = set()

    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str):
                    continue

                urls = URL_PATTERN.findall(cell.value)

                for url in urls:
                    if (
                        "zhipin.com" not in url
                        or JOB_URL_MARKER not in url
                    ):
                        continue

                    try:
                        job_id = get_job_id(url)
                    except ValueError:
                        continue

                    if job_id in seen_job_ids:
                        continue

                    seen_job_ids.add(job_id)

                    records.append(
                        {
                            "job_id": job_id,
                            "source_url": url,
                            "sheet": worksheet.title,
                            "cell": cell.coordinate,
                        }
                    )

    return records


async def get_text(
    page,
    selectors: str | list[str],
    *,
    use_text_content: bool = False,
    multiline: bool = False,
) -> str:
    if isinstance(selectors, str):
        selectors = [selectors]

    for selector in selectors:
        locator = page.locator(selector).first

        try:
            if await locator.count() == 0:
                continue

            if use_text_content:
                value = await locator.text_content(timeout=8_000)
            else:
                value = await locator.inner_text(timeout=8_000)

            if multiline:
                return clean_multiline(value)

            return clean_inline(value)

        except Exception:
            continue

    return ""


async def page_is_ready(page, job_id: str) -> bool:
    if (
        JOB_URL_MARKER not in page.url
        or job_id not in page.url
    ):
        return False

    try:
        await page.wait_for_selector(
            ".boss-info",
            state="attached",
            timeout=15_000,
        )

        await page.wait_for_selector(
            ".rec-position .job-cotent h2 > p",
            state="attached",
            timeout=15_000,
        )

        return True

    except Exception:
        return False


async def open_job_page(
    page,
    url: str,
    job_id: str,
    failure_screenshot: Path,
) -> None:
    for attempt in range(1, 3):
        print(f"  正在访问，第 {attempt} 次尝试……")

        await page.goto(
            url,
            wait_until="domcontentloaded",
            timeout=60_000,
        )

        await page.wait_for_timeout(2_500)

        if await page_is_ready(page, job_id):
            return

        print(f"  当前页面未识别为目标岗位：{page.url}")

        try:
            await page.screenshot(
                path=str(failure_screenshot),
                full_page=True,
            )
        except Exception:
            pass

        if attempt == 1:
            print()
            print("  页面可能出现登录、验证或异常跳转。")
            print("  请在Chrome中手动处理当前页面。")

            await asyncio.to_thread(
                input,
                "  处理完成后按回车重试：",
            )

    raise RuntimeError(
        "两次尝试后仍未进入目标岗位详情页。"
    )


async def extract_job(page, source_record: dict) -> dict:
    source_url = source_record["source_url"]
    expected_job_id = source_record["job_id"]

    final_url = page.url
    final_job_id = get_job_id(final_url)

    if (
        JOB_URL_MARKER not in final_url
        or final_job_id != expected_job_id
    ):
        raise RuntimeError(
            f"页面岗位ID不匹配：{final_url}"
        )

    # 红框完整原文
    core_text = await get_text(
        page,
        ".boss-info",
        multiline=True,
    )

    # 招聘者
    recruiter_name = await get_text(
        page,
        [
            ".boss-info .boss-message .name > p",
            ".boss-message h2.name p",
        ],
    )

    influence = await get_text(
        page,
        [
            ".boss-info .boss-message .name > a",
            ".boss-message h2.name a",
        ],
    )

    recruiter_info_raw = await get_text(
        page,
        ".boss-info .boss-message .info-labels > span",
    )

    recruiter_parts = split_middle_dot(
        recruiter_info_raw
    )

    recruiter_title = (
        recruiter_parts[0]
        if recruiter_parts
        else ""
    )

    recruiter_company = (
        "·".join(recruiter_parts[1:])
        if len(recruiter_parts) >= 2
        else ""
    )

    # 岗位标题和薪资
    job_title = await get_text(
        page,
        [
            ".rec-position .job-cotent h2 > p",
            ".rec-position .job-cotent h2 p",
        ],
    )

    salary = await get_text(
        page,
        ".rec-position .job-cotent > div",
    )

    job_basic_info_raw = await get_text(
        page,
        ".rec-position .job-cotent "
        "h2 > span:not(.job-brandComInfo)",
    )

    basic_parts = split_middle_dot(
        job_basic_info_raw
    )

    city = basic_parts[0] if len(basic_parts) >= 1 else ""
    experience = basic_parts[1] if len(basic_parts) >= 2 else ""
    education = (
        "·".join(basic_parts[2:])
        if len(basic_parts) >= 3
        else ""
    )

    # 职位描述：直接读取DOM，不点击“查看全部”
    job_description = await get_text(
        page,
        ".rec-position .rec-detail .detail-text",
        use_text_content=True,
        multiline=True,
    )

    job_description = re.sub(
        r"(?:\n)?查看全部\s*$",
        "",
        job_description,
    ).strip()

    work_address = await get_text(
        page,
        ".rec-position .detail-text.address",
        multiline=True,
    )

    # 公司信息
    company_full_name = await get_text(
        page,
        ".rec-position .brandComBaseInfo > span",
    )

    company_info_raw = await get_text(
        page,
        ".rec-position .brandComBaseInfo > p",
    )

    company_parts = split_middle_dot(
        company_info_raw
    )

    financing_stage = (
        company_parts[0]
        if len(company_parts) >= 1
        else ""
    )

    company_size = (
        company_parts[1]
        if len(company_parts) >= 2
        else ""
    )

    industry = (
        "·".join(company_parts[2:])
        if len(company_parts) >= 3
        else ""
    )

    company_short_stage_raw = await get_text(
        page,
        ".rec-position .job-brandComInfo",
    )

    # 修复上一版“纳龙健康科技 B轮”没有拆开的情况
    company_short_name = company_short_stage_raw

    if (
        financing_stage
        and company_short_name.endswith(financing_stage)
    ):
        company_short_name = company_short_name[
            :-len(financing_stage)
        ].strip()

    if not company_short_name:
        company_short_name = recruiter_company

    required_fields = {
        "recruiter_name": recruiter_name,
        "job_title": job_title,
        "salary": salary,
        "job_description": job_description,
        "company_full_name": company_full_name,
    }

    missing_fields = [
        field
        for field, value in required_fields.items()
        if not value
    ]

    status = (
        "success"
        if not missing_fields
        else "partial"
    )

    return {
        "job_id": final_job_id,
        "recruiter_name": recruiter_name,
        "influence": influence,
        "recruiter_title": recruiter_title,
        "recruiter_company": recruiter_company,
        "job_title": job_title,
        "salary": salary,
        "city": city,
        "experience": experience,
        "education": education,
        "company_short_name": company_short_name,
        "financing_stage": financing_stage,
        "job_description": job_description,
        "work_address": work_address,
        "company_full_name": company_full_name,
        "company_size": company_size,
        "industry": industry,
        "recruiter_info_raw": recruiter_info_raw,
        "job_basic_info_raw": job_basic_info_raw,
        "company_info_raw": company_info_raw,
        "core_text": core_text,
        "source_url": source_url,
        "final_url": final_url,
        "source_sheet": source_record["sheet"],
        "source_cell": source_record["cell"],
        "collected_at": datetime.now()
        .astimezone()
        .isoformat(),
        "status": status,
        "missing_fields": missing_fields,
    }


def load_existing_records(jsonl_path: Path) -> dict:
    records = {}

    if not jsonl_path.exists():
        return records

    with jsonl_path.open(
        "r",
        encoding="utf-8",
    ) as file:
        for line in file:
            line = line.strip()

            if not line:
                continue

            try:
                record = json.loads(line)
                job_id = record.get("job_id")

                if job_id:
                    records[job_id] = record
            except json.JSONDecodeError:
                continue

    return records


def save_success_outputs(
    records_by_id: dict,
    output_dir: Path,
) -> None:
    records = list(records_by_id.values())

    jsonl_path = output_dir / "jobs.jsonl"
    csv_path = output_dir / "jobs.csv"
    xlsx_path = output_dir / "jobs.xlsx"

    with jsonl_path.open(
        "w",
        encoding="utf-8",
    ) as file:
        for record in records:
            file.write(
                json.dumps(
                    record,
                    ensure_ascii=False,
                )
                + "\n"
            )

    dataframe = pd.DataFrame(records)

    if not dataframe.empty:
        dataframe.to_csv(
            csv_path,
            index=False,
            encoding="utf-8-sig",
        )

        dataframe.to_excel(
            xlsx_path,
            index=False,
            engine="openpyxl",
        )


def save_failures(
    failures: list[dict],
    output_dir: Path,
) -> None:
    if not failures:
        return

    dataframe = pd.DataFrame(failures)

    dataframe.to_csv(
        output_dir / "failed.csv",
        index=False,
        encoding="utf-8-sig",
    )

    dataframe.to_excel(
        output_dir / "failed.xlsx",
        index=False,
        engine="openpyxl",
    )


async def run(args) -> None:
    excel_path = Path(args.input)
    output_dir = Path(args.output)

    if args.force and output_dir.exists():
        shutil.rmtree(output_dir)

    raw_html_dir = output_dir / "raw_pages"
    core_text_dir = output_dir / "core_text"
    failure_screenshot_dir = (
        output_dir / "failed_screenshots"
    )

    raw_html_dir.mkdir(parents=True, exist_ok=True)
    core_text_dir.mkdir(parents=True, exist_ok=True)
    failure_screenshot_dir.mkdir(
        parents=True,
        exist_ok=True,
    )

    source_records = read_job_urls(excel_path)

    if args.start > 1:
        source_records = source_records[
            args.start - 1:
        ]

    if args.limit is not None:
        source_records = source_records[:args.limit]

    if not source_records:
        raise RuntimeError("没有找到可处理的岗位链接。")

    records_by_id = load_existing_records(
        output_dir / "jobs.jsonl"
    )

    failures = []

    print("=" * 72)
    print(f"Excel文件：{excel_path.resolve()}")
    print(f"本次准备处理：{len(source_records)} 条")
    print(f"已有记录：{len(records_by_id)} 条")
    print(f"输出目录：{output_dir.resolve()}")
    print("=" * 72)

    playwright = await async_playwright().start()

    try:
        browser = await playwright.chromium.connect_over_cdp(
            CDP_URL
        )

        if not browser.contexts:
            raise RuntimeError(
                "没有找到可用的Chrome上下文。"
            )

        context = browser.contexts[0]
        page = await context.new_page()

        total = len(source_records)

        for index, source_record in enumerate(
            source_records,
            start=1,
        ):
            job_id = source_record["job_id"]
            source_url = source_record["source_url"]

            existing = records_by_id.get(job_id)

            if (
                not args.force
                and existing
                and existing.get("status") == "success"
            ):
                print(
                    f"[{index}/{total}] 已存在，跳过："
                    f"{existing.get('job_title', job_id)}"
                )
                continue

            print()
            print("-" * 72)
            print(f"[{index}/{total}] 岗位ID：{job_id}")
            print(f"[{index}/{total}] 链接：{source_url}")

            failure_screenshot = (
                failure_screenshot_dir
                / f"{job_id}.png"
            )

            try:
                await open_job_page(
                    page,
                    source_url,
                    job_id,
                    failure_screenshot,
                )

                result = await extract_job(
                    page,
                    source_record,
                )

                html = await page.content()

                (
                    raw_html_dir
                    / f"{job_id}.html"
                ).write_text(
                    html,
                    encoding="utf-8",
                )

                (
                    core_text_dir
                    / f"{job_id}.txt"
                ).write_text(
                    result["core_text"],
                    encoding="utf-8",
                )

                records_by_id[job_id] = result

                save_success_outputs(
                    records_by_id,
                    output_dir,
                )

                print(
                    f"[{index}/{total}] "
                    f"{result['status']}："
                    f"{result['job_title']}"
                )
                print(
                    f"[{index}/{total}] "
                    f"{result['company_full_name']} | "
                    f"{result['salary']}"
                )

                if result["missing_fields"]:
                    print(
                        f"[{index}/{total}] "
                        "缺失字段："
                        + ", ".join(
                            result["missing_fields"]
                        )
                    )

            except KeyboardInterrupt:
                raise

            except Exception as exc:
                failure = {
                    "job_id": job_id,
                    "source_url": source_url,
                    "error_type": type(exc).__name__,
                    "error_message": str(exc),
                    "failed_at": datetime.now()
                    .astimezone()
                    .isoformat(),
                }

                failures.append(failure)
                save_failures(failures, output_dir)

                print(
                    f"[{index}/{total}] 失败："
                    f"{type(exc).__name__}: {exc}"
                )

                try:
                    await page.screenshot(
                        path=str(failure_screenshot),
                        full_page=True,
                    )
                except Exception:
                    pass

            if index < total:
                delay = random.uniform(
                    args.delay_min,
                    args.delay_max,
                )

                print(f"等待 {delay:.1f} 秒……")
                await page.wait_for_timeout(
                    int(delay * 1000)
                )

        await page.close()

    finally:
        await playwright.stop()

    success_count = sum(
        1
        for record in records_by_id.values()
        if record.get("status") == "success"
    )

    partial_count = sum(
        1
        for record in records_by_id.values()
        if record.get("status") == "partial"
    )

    print()
    print("=" * 72)
    print("批量采集完成")
    print("=" * 72)
    print(f"成功：{success_count}")
    print(f"部分成功：{partial_count}")
    print(f"本次失败：{len(failures)}")
    print(f"结果文件：{output_dir / 'jobs.xlsx'}")
    print(f"失败记录：{output_dir / 'failed.xlsx'}")


def parse_args():
    parser = argparse.ArgumentParser(
        description="批量采集BOSS分享岗位"
    )

    parser.add_argument(
        "--input",
        default="BOSS链接.xlsx",
    )

    parser.add_argument(
        "--output",
        default="output/boss_batch",
    )

    parser.add_argument(
        "--start",
        type=int,
        default=1,
        help="从第几条开始，默认1",
    )

    parser.add_argument(
        "--limit",
        type=int,
        default=None,
        help="只处理前N条",
    )

    parser.add_argument(
        "--delay-min",
        type=float,
        default=2.0,
    )

    parser.add_argument(
        "--delay-max",
        type=float,
        default=4.0,
    )

    parser.add_argument(
        "--force",
        action="store_true",
        help="清空旧输出并重新采集",
    )

    return parser.parse_args()


if __name__ == "__main__":
    arguments = parse_args()
    asyncio.run(run(arguments))
