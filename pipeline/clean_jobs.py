import json
import re
from collections import Counter
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


INPUT_FILE = Path("output/boss_batch/jobs.jsonl")
OUTPUT_DIR = Path("output/boss_cleaned")

OUTPUT_JSONL = OUTPUT_DIR / "jobs_cleaned.jsonl"
OUTPUT_CSV = OUTPUT_DIR / "jobs_cleaned.csv"
OUTPUT_XLSX = OUTPUT_DIR / "jobs_cleaned.xlsx"
OUTPUT_AUDIT = OUTPUT_DIR / "field_changes.csv"
OUTPUT_REPORT = OUTPUT_DIR / "quality_report.txt"

CLEANING_VERSION = "1.0"


EDUCATION_PATTERN = re.compile(
    r"^(?:"
    r"学历不限|不限|"
    r"博士|博士及以上|"
    r"硕士|硕士及以上|"
    r"本科|本科及以上|"
    r"大专|大专及以上|"
    r"中专/中技|中专|高中|初中及以下"
    r")$"
)

INTERNSHIP_DAYS_PATTERN = re.compile(
    r"^(?:每周)?\d+(?:-\d+)?天/周$"
)

INTERNSHIP_DURATION_PATTERN = re.compile(
    r"^(?:至少)?\d+(?:-\d+)?个月(?:以上)?$"
)

COMPANY_SIZE_PATTERN = re.compile(
    r"^(?:"
    r"\d+-\d+人|"
    r"\d+人以上|"
    r"\d+人以下|"
    r"\d+人"
    r")$"
)

FINANCING_PATTERN = re.compile(
    r"^(?:"
    r"未融资|不需要融资|融资未公开|"
    r"天使轮|种子轮|"
    r"Pre-A轮|A轮|A\+轮|"
    r"Pre-B轮|B轮|B\+轮|"
    r"C轮|C\+轮|"
    r"D轮|D轮及以上|D\+轮|E轮|"
    r"战略融资|股权融资|定向增发|已上市"
    r")$",
    re.IGNORECASE,
)


def clean_text(value) -> str:
    if value is None:
        return ""

    return re.sub(
        r"\s+",
        " ",
        str(value).replace("\xa0", " "),
    ).strip()


def split_middle_dot(value: str) -> list[str]:
    value = clean_text(value)

    if not value:
        return []

    return [
        part.strip()
        for part in re.split(r"\s*[·•]\s*", value)
        if part.strip()
    ]


def load_jsonl(path: Path) -> list[dict]:
    if not path.exists():
        raise FileNotFoundError(
            f"找不到输入文件：{path.resolve()}\n"
            "请确认批量采集结果位于 output/boss_batch/jobs.jsonl"
        )

    records = []
    seen_job_ids = set()

    with path.open("r", encoding="utf-8") as file:
        for line_number, line in enumerate(file, start=1):
            line = line.strip()

            if not line:
                continue

            try:
                record = json.loads(line)
            except json.JSONDecodeError as exc:
                raise RuntimeError(
                    f"第 {line_number} 行不是有效JSON：{exc}"
                ) from exc

            job_id = clean_text(record.get("job_id"))

            if not job_id:
                raise RuntimeError(
                    f"第 {line_number} 行缺少 job_id"
                )

            if job_id in seen_job_ids:
                raise RuntimeError(
                    f"发现重复岗位ID：{job_id}"
                )

            seen_job_ids.add(job_id)
            records.append(record)

    if not records:
        raise RuntimeError("输入文件中没有岗位记录。")

    return records


def parse_job_basic_info(
    raw_value: str,
    job_title: str,
) -> dict:
    """
    处理两种主要结构：

    正式岗：
    深圳·经验不限·本科

    实习岗：
    深圳·4天/周·6个月·本科
    """
    parts = split_middle_dot(raw_value)

    result = {
        "city": "",
        "experience": "",
        "education": "",
        "internship_days_per_week": "",
        "internship_duration": "",
        "employment_type": "",
    }

    if not parts:
        return result

    result["city"] = parts[0]

    remaining = parts[1:]
    used_indexes = set()

    # 学历通常位于最后，但按模式识别更稳妥
    for index in range(len(remaining) - 1, -1, -1):
        part = remaining[index]

        if EDUCATION_PATTERN.fullmatch(part):
            result["education"] = part
            used_indexes.add(index)
            break

    for index, part in enumerate(remaining):
        if index in used_indexes:
            continue

        if INTERNSHIP_DAYS_PATTERN.fullmatch(part):
            result["internship_days_per_week"] = part
            used_indexes.add(index)
            continue

        if INTERNSHIP_DURATION_PATTERN.fullmatch(part):
            result["internship_duration"] = part
            used_indexes.add(index)

    experience_parts = [
        part
        for index, part in enumerate(remaining)
        if index not in used_indexes
    ]

    result["experience"] = "·".join(experience_parts)

    title = clean_text(job_title)

    if (
        result["internship_days_per_week"]
        or result["internship_duration"]
    ):
        result["employment_type"] = "实习"
    elif "实习" in title and (
        "转正" in title or "可转正" in title
    ):
        result["employment_type"] = "实习/可转正"
    elif "实习" in title:
        result["employment_type"] = "实习"
    elif re.search(
        r"在校/应届|应届生|应届",
        result["experience"],
    ):
        result["employment_type"] = "校招"
    else:
        result["employment_type"] = "常规招聘"

    return result


def parse_company_info(raw_value: str) -> dict:
    """
    支持：

    B轮·500-999人·医疗健康
    20-99人·人工智能
    0-20人·电子商务
    已上市·10000人以上·互联网
    """
    parts = split_middle_dot(raw_value)

    result = {
        "financing_stage": "",
        "company_size": "",
        "industry": "",
    }

    if not parts:
        return result

    size_index = None

    for index, part in enumerate(parts):
        if COMPANY_SIZE_PATTERN.fullmatch(part):
            size_index = index
            result["company_size"] = part
            break

    if size_index is not None:
        before_size = parts[:size_index]
        after_size = parts[size_index + 1:]

        if before_size:
            result["financing_stage"] = "·".join(
                before_size
            )

        if after_size:
            result["industry"] = "·".join(
                after_size
            )

        return result

    # 没识别到规模时，尝试识别融资阶段
    if FINANCING_PATTERN.fullmatch(parts[0]):
        result["financing_stage"] = parts[0]

        if len(parts) >= 2:
            result["industry"] = "·".join(parts[1:])
    else:
        result["industry"] = "·".join(parts)

    return result


def validate_record(record: dict) -> list[str]:
    issues = []

    required_fields = [
        "job_id",
        "job_title",
        "salary",
        "city",
        "education",
        "job_description",
        "company_full_name",
    ]

    for field in required_fields:
        if not clean_text(record.get(field)):
            issues.append(f"缺少字段:{field}")

    education = clean_text(record.get("education"))

    if (
        education
        and not EDUCATION_PATTERN.fullmatch(education)
    ):
        issues.append(f"学历格式异常:{education}")

    days = clean_text(
        record.get("internship_days_per_week")
    )

    if (
        days
        and not INTERNSHIP_DAYS_PATTERN.fullmatch(days)
    ):
        issues.append(f"实习出勤格式异常:{days}")

    duration = clean_text(
        record.get("internship_duration")
    )

    if (
        duration
        and not INTERNSHIP_DURATION_PATTERN.fullmatch(
            duration
        )
    ):
        issues.append(f"实习周期格式异常:{duration}")

    company_size = clean_text(
        record.get("company_size")
    )

    if (
        company_size
        and not COMPANY_SIZE_PATTERN.fullmatch(
            company_size
        )
    ):
        issues.append(f"公司规模格式异常:{company_size}")

    financing_stage = clean_text(
        record.get("financing_stage")
    )

    if (
        financing_stage
        and not FINANCING_PATTERN.fullmatch(
            financing_stage
        )
    ):
        issues.append(
            f"融资阶段待确认:{financing_stage}"
        )

    raw_basic = clean_text(
        record.get("job_basic_info_raw")
    )

    if (
        "天/周" in raw_basic
        and not record.get(
            "internship_days_per_week"
        )
    ):
        issues.append("未识别实习出勤天数")

    if (
        "个月" in raw_basic
        and not record.get("internship_duration")
    ):
        issues.append("未识别实习周期")

    raw_company = clean_text(
        record.get("company_info_raw")
    )

    if (
        re.search(r"\d+(?:-\d+)?人", raw_company)
        and not company_size
    ):
        issues.append("未识别公司规模")

    return issues


def clean_record(record: dict) -> tuple[dict, dict]:
    original = dict(record)
    cleaned = dict(record)

    job_info = parse_job_basic_info(
        clean_text(record.get("job_basic_info_raw")),
        clean_text(record.get("job_title")),
    )

    company_info = parse_company_info(
        clean_text(record.get("company_info_raw"))
    )

    cleaned.update(job_info)
    cleaned.update(company_info)

    cleaned["collection_status"] = clean_text(
        record.get("status")
    )

    cleaned["cleaned_at"] = (
        datetime.now()
        .astimezone()
        .isoformat()
    )

    cleaned["cleaning_version"] = CLEANING_VERSION

    issues = validate_record(cleaned)

    cleaned["quality_status"] = (
        "valid" if not issues else "review"
    )

    cleaned["quality_issue_count"] = len(issues)
    cleaned["quality_issues"] = issues

    audit_fields = [
        "city",
        "experience",
        "education",
        "internship_days_per_week",
        "internship_duration",
        "employment_type",
        "financing_stage",
        "company_size",
        "industry",
    ]

    audit = {
        "job_id": clean_text(record.get("job_id")),
        "job_title": clean_text(
            record.get("job_title")
        ),
    }

    changed_fields = []

    for field in audit_fields:
        before = clean_text(original.get(field))
        after = clean_text(cleaned.get(field))

        audit[f"before_{field}"] = before
        audit[f"after_{field}"] = after

        if before != after:
            changed_fields.append(field)

    audit["changed_fields"] = "；".join(
        changed_fields
    )

    audit["quality_status"] = cleaned[
        "quality_status"
    ]

    audit["quality_issues"] = "；".join(issues)

    return cleaned, audit


def tabular_record(record: dict) -> dict:
    result = {}

    for key, value in record.items():
        if isinstance(value, list):
            result[key] = "；".join(
                clean_text(item)
                for item in value
            )
        elif isinstance(value, dict):
            result[key] = json.dumps(
                value,
                ensure_ascii=False,
            )
        else:
            result[key] = value

    return result


def format_worksheet(worksheet, widths: dict) -> None:
    header_fill = PatternFill(
        fill_type="solid",
        fgColor="0F766E",
    )

    header_font = Font(
        color="FFFFFF",
        bold=True,
    )

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.row_dimensions[1].height = 28

    for column_index, cell in enumerate(
        worksheet[1],
        start=1,
    ):
        column_name = str(cell.value)
        width = widths.get(column_name, 16)

        worksheet.column_dimensions[
            get_column_letter(column_index)
        ].width = width

    for row in worksheet.iter_rows(
        min_row=2,
        max_row=worksheet.max_row,
    ):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )


def save_outputs(
    cleaned_records: list[dict],
    audit_records: list[dict],
) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    with OUTPUT_JSONL.open(
        "w",
        encoding="utf-8",
    ) as file:
        for record in cleaned_records:
            file.write(
                json.dumps(
                    record,
                    ensure_ascii=False,
                )
                + "\n"
            )

    tabular_records = [
        tabular_record(record)
        for record in cleaned_records
    ]

    raw_dataframe = pd.DataFrame(
        tabular_records
    )

    raw_dataframe.to_csv(
        OUTPUT_CSV,
        index=False,
        encoding="utf-8-sig",
    )

    summary_columns = [
        "job_id",
        "job_title",
        "company_full_name",
        "company_short_name",
        "salary",
        "city",
        "employment_type",
        "internship_days_per_week",
        "internship_duration",
        "experience",
        "education",
        "financing_stage",
        "company_size",
        "industry",
        "recruiter_name",
        "recruiter_title",
        "quality_status",
        "quality_issue_count",
        "quality_issues",
        "source_url",
    ]

    for column in summary_columns:
        if column not in raw_dataframe.columns:
            raw_dataframe[column] = ""

    summary_dataframe = raw_dataframe[
        summary_columns
    ].copy()

    review_dataframe = raw_dataframe[
        raw_dataframe["quality_status"] == "review"
    ].copy()

    audit_dataframe = pd.DataFrame(
        audit_records
    )

    audit_dataframe.to_csv(
        OUTPUT_AUDIT,
        index=False,
        encoding="utf-8-sig",
    )

    with pd.ExcelWriter(
        OUTPUT_XLSX,
        engine="openpyxl",
    ) as writer:
        summary_dataframe.to_excel(
            writer,
            sheet_name="Jobs_Summary",
            index=False,
        )

        raw_dataframe.to_excel(
            writer,
            sheet_name="Jobs_Raw",
            index=False,
        )

        audit_dataframe.to_excel(
            writer,
            sheet_name="Field_Changes",
            index=False,
        )

        if review_dataframe.empty:
            pd.DataFrame(
                {
                    "检查结果": [
                        "所有岗位均通过字段规则检查"
                    ]
                }
            ).to_excel(
                writer,
                sheet_name="Needs_Review",
                index=False,
            )
        else:
            review_dataframe.to_excel(
                writer,
                sheet_name="Needs_Review",
                index=False,
            )

        workbook = writer.book

        summary_widths = {
            "job_id": 26,
            "job_title": 32,
            "company_full_name": 28,
            "company_short_name": 22,
            "salary": 14,
            "city": 10,
            "employment_type": 14,
            "internship_days_per_week": 16,
            "internship_duration": 14,
            "experience": 15,
            "education": 14,
            "financing_stage": 14,
            "company_size": 14,
            "industry": 18,
            "recruiter_name": 12,
            "recruiter_title": 18,
            "quality_status": 14,
            "quality_issue_count": 12,
            "quality_issues": 34,
            "source_url": 35,
        }

        raw_widths = {
            column: 18
            for column in raw_dataframe.columns
        }

        for long_column in [
            "job_title",
            "company_full_name",
            "job_description",
            "work_address",
            "core_text",
            "source_url",
            "final_url",
            "quality_issues",
        ]:
            raw_widths[long_column] = 40

        format_worksheet(
            workbook["Jobs_Summary"],
            summary_widths,
        )

        format_worksheet(
            workbook["Jobs_Raw"],
            raw_widths,
        )

        format_worksheet(
            workbook["Field_Changes"],
            {
                column: 20
                for column in audit_dataframe.columns
            },
        )

        format_worksheet(
            workbook["Needs_Review"],
            {
                column: 28
                for column in workbook[
                    "Needs_Review"
                ][1]
            },
        )


def write_quality_report(
    cleaned_records: list[dict],
) -> None:
    quality_counts = Counter(
        record["quality_status"]
        for record in cleaned_records
    )

    employment_counts = Counter(
        record["employment_type"]
        for record in cleaned_records
    )

    review_records = [
        record
        for record in cleaned_records
        if record["quality_status"] == "review"
    ]

    lines = [
        "=" * 60,
        "BOSS岗位数据离线清洗报告",
        "=" * 60,
        f"记录总数：{len(cleaned_records)}",
        f"通过检查：{quality_counts.get('valid', 0)}",
        f"需要复核：{quality_counts.get('review', 0)}",
        "",
        "招聘类型：",
    ]

    for name, count in sorted(
        employment_counts.items()
    ):
        lines.append(f"- {name or '未分类'}：{count}")

    if review_records:
        lines.extend(
            [
                "",
                "需要复核的岗位：",
            ]
        )

        for record in review_records:
            issues = "；".join(
                record["quality_issues"]
            )

            lines.append(
                f"- {record.get('job_title', '')}"
                f" | {issues}"
            )

    OUTPUT_REPORT.write_text(
        "\n".join(lines),
        encoding="utf-8",
    )


def main() -> None:
    records = load_jsonl(INPUT_FILE)

    cleaned_records = []
    audit_records = []

    for record in records:
        cleaned, audit = clean_record(record)
        cleaned_records.append(cleaned)
        audit_records.append(audit)

    save_outputs(
        cleaned_records,
        audit_records,
    )

    write_quality_report(cleaned_records)

    valid_count = sum(
        record["quality_status"] == "valid"
        for record in cleaned_records
    )

    review_count = len(cleaned_records) - valid_count

    internship_count = sum(
        record["employment_type"]
        in {"实习", "实习/可转正"}
        for record in cleaned_records
    )

    print("=" * 70)
    print("离线清洗完成")
    print("=" * 70)
    print(f"输入记录：{len(records)}")
    print(f"实习相关岗位：{internship_count}")
    print(f"通过规则检查：{valid_count}")
    print(f"需要人工复核：{review_count}")
    print()
    print(f"清洗后JSONL：{OUTPUT_JSONL.resolve()}")
    print(f"清洗后CSV：{OUTPUT_CSV.resolve()}")
    print(f"清洗后Excel：{OUTPUT_XLSX.resolve()}")
    print(f"字段变化记录：{OUTPUT_AUDIT.resolve()}")
    print(f"质量报告：{OUTPUT_REPORT.resolve()}")


if __name__ == "__main__":
    main()
