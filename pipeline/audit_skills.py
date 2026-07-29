import json
import re
from collections import Counter, defaultdict
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


INPUT_FILE = Path(
    "output/boss_cleaned/jobs_cleaned.jsonl"
)

OLD_FREQUENCY_FILE = Path(
    "output/analysis_v1/skill_frequency.csv"
)

OUTPUT_DIR = Path("output/analysis_v1_1")
ANALYSIS_VERSION = "1.1"


# ============================================================
# 技能词典
# 每项技能同时带有所属维度和匹配规则
# ============================================================

SKILL_DEFS = [
    # ---------------- 技术栈 ----------------
    (
        "Python",
        "技术栈",
        [
            r"(?<![a-z0-9])python(?![a-z0-9])",
        ],
    ),
    (
        "SQL",
        "技术栈",
        [
            r"(?<![a-z0-9])sql(?![a-z0-9])",
        ],
    ),
    (
        "JavaScript",
        "技术栈",
        [
            r"(?<![a-z0-9])javascript(?![a-z0-9])",
            r"(?<![a-z0-9])js(?![a-z0-9])",
        ],
    ),
    (
        "Java",
        "技术栈",
        [
            r"(?<![a-z0-9])java(?![a-z0-9])",
        ],
    ),
    (
        "C++",
        "技术栈",
        [
            r"c\+\+",
        ],
    ),
    (
        "C语言",
        "技术栈",
        [
            # 不再匹配孤立字母C，避免把“C座”当成C语言
            r"c语言",
            r"c\s*/\s*c\+\+",
            r"熟悉\s*c(?:编程|开发|语言)",
        ],
    ),
    (
        "R语言",
        "技术栈",
        [
            r"r语言",
            r"(?<![a-z0-9])r(?![a-z0-9])"
            r"(?=\s*(?:、|/|和|与|语言|\)|）|,|，))",
        ],
    ),
    (
        "Pandas",
        "技术栈",
        [
            r"(?<![a-z0-9])pandas(?![a-z0-9])",
        ],
    ),
    (
        "NumPy",
        "技术栈",
        [
            r"(?<![a-z0-9])numpy(?![a-z0-9])",
        ],
    ),
    (
        "Scikit-learn",
        "技术栈",
        [
            r"scikit[- ]?learn",
            r"(?<![a-z0-9])sklearn(?![a-z0-9])",
        ],
    ),
    (
        "Matplotlib",
        "技术栈",
        [
            r"(?<![a-z0-9])matplotlib(?![a-z0-9])",
        ],
    ),
    (
        "PyTorch",
        "技术栈",
        [
            r"(?<![a-z0-9])pytorch(?![a-z0-9])",
            r"(?<![a-z0-9])torch(?![a-z0-9])",
        ],
    ),
    (
        "TensorFlow",
        "技术栈",
        [
            r"(?<![a-z0-9])tensorflow(?![a-z0-9])",
        ],
    ),
    (
        "Excel",
        "技术栈",
        [
            r"(?<![a-z0-9])excel(?![a-z0-9])",
        ],
    ),
    (
        "Power BI",
        "技术栈",
        [
            r"power\s*bi",
            r"powerbi",
        ],
    ),
    (
        "Tableau",
        "技术栈",
        [
            r"(?<![a-z0-9])tableau(?![a-z0-9])",
        ],
    ),
    (
        "FineBI",
        "技术栈",
        [
            r"(?<![a-z0-9])finebi(?![a-z0-9])",
        ],
    ),
    (
        "Linux",
        "技术栈",
        [
            r"(?<![a-z0-9])linux(?![a-z0-9])",
        ],
    ),
    (
        "Docker",
        "技术栈",
        [
            r"(?<![a-z0-9])docker(?![a-z0-9])",
        ],
    ),
    (
        "Git",
        "技术栈",
        [
            r"(?<![a-z0-9])git(?![a-z0-9])",
        ],
    ),
    (
        "Spark",
        "技术栈",
        [
            r"(?<![a-z0-9])spark(?![a-z0-9])",
        ],
    ),
    (
        "Hadoop",
        "技术栈",
        [
            r"(?<![a-z0-9])hadoop(?![a-z0-9])",
        ],
    ),
    (
        "Hive",
        "技术栈",
        [
            r"(?<![a-z0-9])hive(?![a-z0-9])",
        ],
    ),
    (
        "API/接口",
        "技术栈",
        [
            r"(?<![a-z0-9])api(?![a-z0-9])",
            r"接口设计",
            r"接口开发",
        ],
    ),
    (
        "正则表达式",
        "技术栈",
        [
            r"正则(?:表达式|处理)?",
            r"regular expression",
        ],
    ),
    (
        "爬虫",
        "技术栈",
        [
            r"爬虫",
            r"网页抓取",
            r"网络抓取",
        ],
    ),

    # ---------------- AI与统计方法 ----------------
    (
        "机器学习",
        "AI与统计方法",
        [
            r"机器学习",
            r"machine learning",
        ],
    ),
    (
        "深度学习",
        "AI与统计方法",
        [
            r"深度学习",
            r"deep learning",
        ],
    ),
    (
        "大模型/LLM",
        "AI与统计方法",
        [
            r"大模型",
            r"语言模型",
            r"(?<![a-z0-9])llm(?![a-z0-9])",
        ],
    ),
    (
        "AI Agent",
        "AI与统计方法",
        [
            r"ai\s*agent",
            r"智能体",
            r"(?<![a-z0-9])agent(?![a-z0-9])",
        ],
    ),
    (
        "RAG",
        "AI与统计方法",
        [
            r"(?<![a-z0-9])rag(?![a-z0-9])",
            r"检索增强生成",
        ],
    ),
    (
        "Prompt",
        "AI与统计方法",
        [
            r"(?<![a-z0-9])prompt(?![a-z0-9])",
            r"提示词",
        ],
    ),
    (
        "Fine-tuning",
        "AI与统计方法",
        [
            r"fine[- ]?tun(?:e|ing)",
            r"模型微调",
            r"大模型微调",
        ],
    ),
    (
        "向量数据库",
        "AI与统计方法",
        [
            r"向量数据库",
            r"向量检索",
            r"vector database",
        ],
    ),
    (
        "知识库",
        "AI与统计方法",
        [
            r"知识库",
            r"知识语料",
        ],
    ),
    (
        "意图识别",
        "AI与统计方法",
        [
            r"意图识别",
            r"(?<![a-z0-9])intent(?![a-z0-9])",
        ],
    ),
    (
        "多轮对话",
        "AI与统计方法",
        [
            r"多轮对话",
        ],
    ),
    (
        "时间序列",
        "AI与统计方法",
        [
            r"时间序列",
            r"时序数据",
            r"高频数据",
        ],
    ),
    (
        "统计分析",
        "AI与统计方法",
        [
            r"统计分析",
            r"统计学",
            r"统计方法",
        ],
    ),
    (
        "A/B测试",
        "AI与统计方法",
        [
            r"a\s*/?\s*b\s*测试",
            r"ab测试",
        ],
    ),
    (
        "假设检验",
        "AI与统计方法",
        [
            r"假设检验",
        ],
    ),
    (
        "量化分析",
        "AI与统计方法",
        [
            r"量化分析",
            r"量化模型",
            r"因子挖掘",
            r"市场微观结构",
        ],
    ),

    # ---------------- 工作任务 ----------------
    (
        "数据分析",
        "工作任务",
        [
            r"数据分析",
            r"业务分析",
            r"分析报告",
        ],
    ),
    (
        "数据可视化",
        "工作任务",
        [
            r"数据可视化",
            r"可视化分析",
            r"可视化图表",
            r"数据看板",
        ],
    ),
    (
        "数据采集",
        "工作任务",
        [
            r"数据采集",
            r"数据收集",
            r"采集数据",
            r"抓取数据",
        ],
    ),
    (
        "数据清洗",
        "工作任务",
        [
            r"数据清洗",
            r"清洗数据",
            r"清洗与过滤",
        ],
    ),
    (
        "数据预处理",
        "工作任务",
        [
            r"数据预处理",
            r"预处理数据",
        ],
    ),
    (
        "数据标注",
        "工作任务",
        [
            r"数据标注",
            r"标注数据",
            r"数据标签",
        ],
    ),
    (
        "数据治理",
        "工作任务",
        [
            r"数据治理",
            r"数据标准化",
            r"数据质量",
            r"主数据管理",
        ],
    ),
    (
        "数据建模",
        "工作任务",
        [
            r"数据建模",
            r"数据模型",
        ],
    ),
    (
        "数据仓库",
        "工作任务",
        [
            r"数据仓库",
            r"数仓",
        ],
    ),
    (
        "ETL",
        "工作任务",
        [
            r"(?<![a-z0-9])etl(?![a-z0-9])",
        ],
    ),
    (
        "模型训练",
        "工作任务",
        [
            r"模型训练",
            r"训练模型",
        ],
    ),
    (
        "模型部署",
        "工作任务",
        [
            r"模型部署",
            r"部署模型",
            r"推理部署",
        ],
    ),
    (
        "模型评测",
        "工作任务",
        [
            r"模型评测",
            r"效果评测",
            r"评测脚本",
            r"模型测试",
        ],
    ),
    (
        "工具调用",
        "工作任务",
        [
            r"工具调用",
            r"tool\s*calling",
            r"skill\s*/?\s*tool",
        ],
    ),
    (
        "工作流编排",
        "工作任务",
        [
            r"工作流",
            r"agent编排",
            r"流程编排",
        ],
    ),
    (
        "自动化脚本",
        "工作任务",
        [
            r"自动化脚本",
            r"脚本开发",
        ],
    ),
    (
        "埋点分析",
        "工作任务",
        [
            r"埋点",
            r"神策",
            r"数数平台",
        ],
    ),
    (
        "业务指标",
        "工作任务",
        [
            r"业务指标",
            r"指标体系",
        ],
    ),
    (
        "异常分析",
        "工作任务",
        [
            r"异常分析",
            r"异常检测",
            r"异常排查",
            r"下钻分析",
        ],
    ),

    # ---------------- 产品与项目 ----------------
    (
        "需求分析",
        "产品与项目",
        [
            r"需求分析",
            r"需求拆解",
            r"业务需求",
        ],
    ),
    (
        "产品设计",
        "产品与项目",
        [
            r"产品设计",
            r"产品定义",
            r"方案设计",
        ],
    ),
    (
        "项目管理",
        "产品与项目",
        [
            r"项目管理",
            r"项目推进",
            r"项目落地",
        ],
    ),

    # ---------------- 通用能力 ----------------
    (
        "跨团队协作",
        "通用能力",
        [
            r"跨团队",
            r"多团队协作",
            r"团队协作",
            r"沟通协调",
        ],
    ),
    (
        "报告撰写",
        "通用能力",
        [
            r"报告撰写",
            r"撰写报告",
            r"分析报告",
            r"工作汇报",
        ],
    ),
    (
        "英语",
        "通用能力",
        [
            r"英语",
            r"英文",
        ],
    ),
]


# ============================================================
# 岗位类别规则
# 标题命中权重为3，正文命中权重为1
# ============================================================

ROLE_DEFS = {
    "数据分析与BI": {
        "title": [
            r"数据分析",
            r"分析师",
            r"数据应用",
            r"大数据分析",
        ],
        "description": [
            r"业务分析",
            r"报表",
            r"看板",
            r"可视化",
            r"统计分析",
            r"指标体系",
        ],
    },
    "AI与大模型开发": {
        "title": [
            r"agent",
            r"大模型开发",
            r"python开发.*agent",
            r"ai工程师",
        ],
        "description": [
            r"大模型",
            r"智能体",
            r"\brag\b",
            r"prompt",
            r"工具调用",
            r"知识库",
        ],
    },
    "算法与机器学习": {
        "title": [
            r"算法",
            r"深度学习",
        ],
        "description": [
            r"机器学习",
            r"深度学习",
            r"模型训练",
            r"pytorch",
            r"tensorflow",
        ],
    },
    "AI产品与项目": {
        "title": [
            r"产品经理",
            r"ai项目",
            r"项目实习生",
        ],
        "description": [
            r"需求分析",
            r"产品设计",
            r"项目推进",
            r"项目管理",
        ],
    },
    "数据处理与标注": {
        "title": [
            r"大模型数据",
            r"数据处理",
            r"流程优化",
            r"人工智能数据分析",
        ],
        "description": [
            r"数据清洗",
            r"数据标注",
            r"数据处理",
            r"语料",
        ],
    },
    "数据工程与平台": {
        "title": [
            r"数据开发",
            r"数据仓库",
            r"数仓",
            r"大数据工程",
        ],
        "description": [
            r"\betl\b",
            r"\bspark\b",
            r"\bhadoop\b",
            r"\bhive\b",
            r"数据平台",
            r"数据仓库",
        ],
    },
    "数据采集与测试": {
        "title": [
            r"数据采集",
            r"测试数据",
        ],
        "description": [
            r"数据采集",
            r"测试数据",
            r"采集设备",
        ],
    },
}


# 疑似招聘文案编辑痕迹
EDITOR_TRACE_PATTERNS = [
    r"\[你的邮箱\]",
    r"请将简历发送至",
    r"更新说明",
    r"新增/强化内容",
    r"以下为.*优化",
    r"优化后的岗位",
]


# 证据强度提示词
REQUIREMENT_CUES = [
    r"必须",
    r"要求",
    r"熟练",
    r"熟悉",
    r"精通",
    r"掌握",
    r"具备",
    r"能够",
    r"需要",
    r"任职资格",
    r"任职要求",
]

PREFERRED_CUES = [
    r"优先",
    r"加分",
    r"更佳",
    r"者优",
]

DUTY_CUES = [
    r"负责",
    r"参与",
    r"协助",
    r"支持",
    r"工作内容",
    r"岗位职责",
    r"完成",
    r"开展",
    r"搭建",
    r"开发",
    r"维护",
    r"推进",
    r"输出",
]

LEVEL_PRIORITY = {
    "必备": 4,
    "优先": 3,
    "职责": 2,
    "普通提及": 1,
    "标题提及": 1,
}


def clean_text(value) -> str:
    if value is None:
        return ""

    return re.sub(
        r"\s+",
        " ",
        str(value).replace("\xa0", " "),
    ).strip()


def load_jsonl(path: Path) -> list[dict]:
    if not path.exists():
        raise FileNotFoundError(
            f"找不到输入文件：{path.resolve()}"
        )

    records = []

    with path.open(
        "r",
        encoding="utf-8",
    ) as file:
        for line_number, line in enumerate(
            file,
            start=1,
        ):
            line = line.strip()

            if not line:
                continue

            try:
                records.append(json.loads(line))
            except json.JSONDecodeError as exc:
                raise RuntimeError(
                    f"第 {line_number} 行JSON解析失败："
                    f"{exc}"
                ) from exc

    if not records:
        raise RuntimeError("输入文件中没有岗位记录。")

    return records


def split_evidence_units(
    title: str,
    description: str,
) -> list[tuple[str, str]]:
    """
    把标题和职位描述拆成可核查的证据单元。
    """
    units = []

    title = clean_text(title)

    if title:
        units.append(("标题", title))

    raw = str(description or "")
    raw = raw.replace("\r\n", "\n").replace("\r", "\n")

    chunks = re.split(
        r"\n+|(?<=[。！？；;])",
        raw,
    )

    for chunk in chunks:
        chunk = clean_text(chunk)

        if chunk:
            units.append(("正文", chunk))

    return units


def classify_evidence(
    source: str,
    sentence: str,
) -> str:
    """
    判断技能在当前句子中属于：
    标题提及、必备、优先、职责或普通提及。
    """
    if source == "标题":
        return "标题提及"

    if any(
        re.search(
            pattern,
            sentence,
            re.IGNORECASE,
        )
        for pattern in PREFERRED_CUES
    ):
        return "优先"

    if any(
        re.search(
            pattern,
            sentence,
            re.IGNORECASE,
        )
        for pattern in REQUIREMENT_CUES
    ):
        return "必备"

    if any(
        re.search(
            pattern,
            sentence,
            re.IGNORECASE,
        )
        for pattern in DUTY_CUES
    ):
        return "职责"

    return "普通提及"


def extract_skills_with_evidence(
    title: str,
    description: str,
) -> tuple[list[dict], list[dict]]:
    units = split_evidence_units(
        title,
        description,
    )

    skill_rows = []
    evidence_rows = []

    for skill_name, group, patterns in SKILL_DEFS:
        hits = []
        seen = set()

        for source, sentence in units:
            matched = any(
                re.search(
                    pattern,
                    sentence,
                    re.IGNORECASE,
                )
                for pattern in patterns
            )

            if not matched:
                continue

            key = (source, sentence)

            if key in seen:
                continue

            seen.add(key)

            level = classify_evidence(
                source,
                sentence,
            )

            hits.append(
                {
                    "source": source,
                    "level": level,
                    "evidence": sentence,
                }
            )

        if not hits:
            continue

        # 每个技能最多保存5条命中证据
        hits = hits[:5]

        aggregate_level = max(
            hits,
            key=lambda item: LEVEL_PRIORITY[
                item["level"]
            ],
        )["level"]

        skill_rows.append(
            {
                "skill": skill_name,
                "group": group,
                "level": aggregate_level,
                "evidence_count": len(hits),
                "evidence": [
                    item["evidence"]
                    for item in hits
                ],
            }
        )

        for hit in hits:
            evidence_rows.append(
                {
                    "skill": skill_name,
                    "skill_group": group,
                    "evidence_source": hit["source"],
                    "requirement_level": hit["level"],
                    "evidence_text": hit["evidence"],
                }
            )

    return skill_rows, evidence_rows


def classify_role(
    title: str,
    description: str,
) -> dict:
    title_lower = clean_text(title).lower()
    description_lower = clean_text(description).lower()

    scores = {}

    for role, rules in ROLE_DEFS.items():
        score = 0
        title_hits = []
        description_hits = []

        for pattern in rules["title"]:
            if re.search(
                pattern,
                title_lower,
                re.IGNORECASE,
            ):
                score += 3
                title_hits.append(pattern)

        for pattern in rules["description"]:
            if re.search(
                pattern,
                description_lower,
                re.IGNORECASE,
            ):
                score += 1
                description_hits.append(pattern)

        scores[role] = {
            "score": score,
            "title_hits": title_hits,
            "description_hits": description_hits,
        }

    ranked = sorted(
        scores.items(),
        key=lambda item: item[1]["score"],
        reverse=True,
    )

    primary, primary_detail = ranked[0]

    second_score = (
        ranked[1][1]["score"]
        if len(ranked) > 1
        else 0
    )

    top_score = primary_detail["score"]

    if top_score == 0:
        primary = "其他/综合"
        confidence = "低"

    elif (
        top_score >= 5
        and top_score - second_score >= 3
    ):
        confidence = "高"

    elif (
        top_score >= 3
        and top_score - second_score >= 1
    ):
        confidence = "中"

    else:
        confidence = "低"

    secondary = [
        role
        for role, detail in ranked[1:]
        if detail["score"] >= 2
    ]

    generic_title = bool(
        re.search(
            r"数据方向|工程师助理|应届生/实习生",
            title_lower,
        )
    )

    review_needed = (
        confidence == "低"
        or generic_title
        or (
            top_score == second_score
            and top_score > 0
        )
    )

    return {
        "role_category_v11": primary,
        "role_secondary_tags": secondary,
        "role_confidence": confidence,
        "role_review_needed": review_needed,
        "role_scores": {
            role: detail["score"]
            for role, detail in scores.items()
        },
    }


def detect_anomalies(
    description: str,
    skill_count: int,
) -> dict:
    description = str(description or "")
    compact = clean_text(description)

    flags = []

    for pattern in EDITOR_TRACE_PATTERNS:
        if re.search(
            pattern,
            compact,
            re.IGNORECASE,
        ):
            flags.append("疑似编辑模板残留")
            break

    if len(compact) >= 1800:
        flags.append("超长JD")

    if len(compact) < 120:
        flags.append("短JD")

    if skill_count <= 1:
        flags.append("明确技能信息较少")

    if (
        len(compact) >= 500
        and skill_count >= 8
    ):
        density = "高"

    elif (
        len(compact) >= 180
        and skill_count >= 3
    ):
        density = "中"

    else:
        density = "低"

    # 核心样本只排除明显带编辑模板残留的记录
    core_sample = (
        "疑似编辑模板残留" not in flags
    )

    return {
        "jd_char_count": len(compact),
        "jd_information_density": density,
        "anomaly_flags": flags,
        "core_sample": core_sample,
    }


def audit_record(
    record: dict,
) -> tuple[dict, list[dict]]:
    result = dict(record)

    title = clean_text(
        record.get("job_title")
    )

    description = str(
        record.get("job_description") or ""
    )

    skill_rows, evidence_rows = (
        extract_skills_with_evidence(
            title,
            description,
        )
    )

    role_info = classify_role(
        title,
        description,
    )

    anomaly_info = detect_anomalies(
        description,
        len(skill_rows),
    )

    result.update(role_info)
    result.update(anomaly_info)

    result["skills_v11"] = [
        item["skill"]
        for item in skill_rows
    ]

    result["skill_groups_v11"] = sorted(
        set(
            item["group"]
            for item in skill_rows
        )
    )

    result["skill_count_v11"] = len(skill_rows)
    result["skill_details_v11"] = skill_rows
    result["analysis_version"] = ANALYSIS_VERSION

    for row in evidence_rows:
        row.update(
            {
                "job_id": record.get(
                    "job_id",
                    "",
                ),
                "job_title": title,
                "company_full_name": record.get(
                    "company_full_name",
                    "",
                ),
                "core_sample": anomaly_info[
                    "core_sample"
                ],
            }
        )

    return result, evidence_rows


def build_frequency(
    records: list[dict],
    core_only: bool,
) -> pd.DataFrame:
    eligible = [
        record
        for record in records
        if (
            record.get("core_sample")
            or not core_only
        )
    ]

    counts = Counter()
    groups = {}
    levels = defaultdict(Counter)

    for record in eligible:
        for detail in record.get(
            "skill_details_v11",
            [],
        ):
            skill = detail["skill"]

            counts[skill] += 1
            groups[skill] = detail["group"]
            levels[skill][detail["level"]] += 1

    rows = []
    total = len(eligible)

    for rank, (skill, count) in enumerate(
        counts.most_common(),
        start=1,
    ):
        rows.append(
            {
                "排名": rank,
                "技能": skill,
                "技能组": groups.get(
                    skill,
                    "",
                ),
                "出现岗位数": count,
                "岗位覆盖率": (
                    round(count / total, 4)
                    if total
                    else 0
                ),
                "岗位覆盖率百分比": (
                    f"{count / total:.1%}"
                    if total
                    else "0.0%"
                ),
                "必备岗位数": levels[
                    skill
                ].get("必备", 0),
                "优先岗位数": levels[
                    skill
                ].get("优先", 0),
                "职责岗位数": levels[
                    skill
                ].get("职责", 0),
                "普通或标题提及岗位数": (
                    levels[skill].get(
                        "普通提及",
                        0,
                    )
                    + levels[skill].get(
                        "标题提及",
                        0,
                    )
                ),
                "样本类型": (
                    "核心样本"
                    if core_only
                    else "全样本"
                ),
                "样本岗位数": total,
            }
        )

    return pd.DataFrame(rows)


def build_group_frequency(
    records: list[dict],
    core_only: bool,
) -> pd.DataFrame:
    eligible = [
        record
        for record in records
        if (
            record.get("core_sample")
            or not core_only
        )
    ]

    counts = Counter()

    for record in eligible:
        groups = set(
            detail["group"]
            for detail in record.get(
                "skill_details_v11",
                [],
            )
        )

        counts.update(groups)

    total = len(eligible)

    return pd.DataFrame(
        [
            {
                "技能组": group,
                "覆盖岗位数": count,
                "岗位覆盖率": (
                    round(count / total, 4)
                    if total
                    else 0
                ),
                "岗位覆盖率百分比": (
                    f"{count / total:.1%}"
                    if total
                    else "0.0%"
                ),
                "样本类型": (
                    "核心样本"
                    if core_only
                    else "全样本"
                ),
            }
            for group, count
            in counts.most_common()
        ]
    )


def tabularize(
    records: list[dict],
) -> pd.DataFrame:
    rows = []

    for record in records:
        rows.append(
            {
                "job_id": record.get(
                    "job_id",
                    "",
                ),
                "job_title": record.get(
                    "job_title",
                    "",
                ),
                "company_full_name": record.get(
                    "company_full_name",
                    "",
                ),
                "city": record.get(
                    "city",
                    "",
                ),
                "salary": record.get(
                    "salary",
                    "",
                ),
                "employment_type": record.get(
                    "employment_type",
                    "",
                ),
                "role_category_v1": record.get(
                    "role_category",
                    "",
                ),
                "role_category_v11": record.get(
                    "role_category_v11",
                    "",
                ),
                "role_secondary_tags": "；".join(
                    record.get(
                        "role_secondary_tags",
                        [],
                    )
                ),
                "role_confidence": record.get(
                    "role_confidence",
                    "",
                ),
                "role_review_needed": record.get(
                    "role_review_needed",
                    False,
                ),
                "jd_char_count": record.get(
                    "jd_char_count",
                    0,
                ),
                "jd_information_density": record.get(
                    "jd_information_density",
                    "",
                ),
                "anomaly_flags": "；".join(
                    record.get(
                        "anomaly_flags",
                        [],
                    )
                ),
                "core_sample": record.get(
                    "core_sample",
                    True,
                ),
                "skill_count_v1": record.get(
                    "skill_count",
                    "",
                ),
                "skill_count_v11": record.get(
                    "skill_count_v11",
                    0,
                ),
                "skills_v11": "；".join(
                    record.get(
                        "skills_v11",
                        [],
                    )
                ),
                "skill_groups_v11": "；".join(
                    record.get(
                        "skill_groups_v11",
                        [],
                    )
                ),
                "source_url": record.get(
                    "source_url",
                    "",
                ),
            }
        )

    return pd.DataFrame(rows)


def build_frequency_comparison(
    new_frequency: pd.DataFrame,
) -> pd.DataFrame:
    if not OLD_FREQUENCY_FILE.exists():
        return pd.DataFrame()

    old = pd.read_csv(
        OLD_FREQUENCY_FILE,
        encoding="utf-8-sig",
    )

    old = old[
        [
            "技能",
            "出现岗位数",
        ]
    ].rename(
        columns={
            "出现岗位数": "v1出现岗位数",
        }
    )

    new = new_frequency[
        [
            "技能",
            "出现岗位数",
        ]
    ].rename(
        columns={
            "出现岗位数": "v1.1出现岗位数",
        }
    )

    merged = old.merge(
        new,
        on="技能",
        how="outer",
    ).fillna(0)

    merged["v1出现岗位数"] = (
        merged["v1出现岗位数"].astype(int)
    )

    merged["v1.1出现岗位数"] = (
        merged["v1.1出现岗位数"].astype(int)
    )

    merged["变化"] = (
        merged["v1.1出现岗位数"]
        - merged["v1出现岗位数"]
    )

    return merged.sort_values(
        [
            "变化",
            "v1.1出现岗位数",
        ],
        ascending=[
            True,
            False,
        ],
    )


def format_sheet(
    worksheet,
    widths: dict | None = None,
) -> None:
    widths = widths or {}

    fill = PatternFill(
        fill_type="solid",
        fgColor="0F766E",
    )

    font = Font(
        color="FFFFFF",
        bold=True,
    )

    for cell in worksheet[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for index, cell in enumerate(
        worksheet[1],
        start=1,
    ):
        column_name = str(cell.value)

        worksheet.column_dimensions[
            get_column_letter(index)
        ].width = widths.get(
            column_name,
            18,
        )

    for row in worksheet.iter_rows(
        min_row=2,
    ):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )


def markdown_table(
    dataframe: pd.DataFrame,
    columns: list[str],
    max_rows: int = 15,
) -> str:
    if dataframe.empty:
        return "暂无数据。"

    frame = dataframe[
        columns
    ].head(max_rows).copy()

    lines = [
        "| " + " | ".join(columns) + " |",
        "| "
        + " | ".join(
            ["---"] * len(columns)
        )
        + " |",
    ]

    for _, row in frame.iterrows():
        values = [
            clean_text(value).replace(
                "|",
                "\\|",
            )
            for value in row.tolist()
        ]

        lines.append(
            "| " + " | ".join(values) + " |"
        )

    return "\n".join(lines)


def main() -> None:
    OUTPUT_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    records = load_jsonl(INPUT_FILE)

    audited_records = []
    evidence_rows = []

    for record in records:
        audited, evidence = audit_record(record)

        audited_records.append(audited)
        evidence_rows.extend(evidence)

    full_frequency = build_frequency(
        audited_records,
        core_only=False,
    )

    core_frequency = build_frequency(
        audited_records,
        core_only=True,
    )

    full_groups = build_group_frequency(
        audited_records,
        core_only=False,
    )

    core_groups = build_group_frequency(
        audited_records,
        core_only=True,
    )

    jobs_dataframe = tabularize(
        audited_records
    )

    evidence_dataframe = pd.DataFrame(
        evidence_rows
    )

    role_review_dataframe = jobs_dataframe[
        jobs_dataframe[
            "role_review_needed"
        ] == True
    ].copy()

    anomaly_dataframe = jobs_dataframe[
        (
            jobs_dataframe[
                "anomaly_flags"
            ] != ""
        )
        | (
            jobs_dataframe[
                "jd_information_density"
            ] == "低"
        )
    ].copy()

    comparison_dataframe = (
        build_frequency_comparison(
            full_frequency
        )
    )

    # 保存带完整证据结构的JSONL
    with (
        OUTPUT_DIR
        / "jobs_skill_audited.jsonl"
    ).open(
        "w",
        encoding="utf-8",
    ) as file:
        for record in audited_records:
            file.write(
                json.dumps(
                    record,
                    ensure_ascii=False,
                )
                + "\n"
            )

    evidence_dataframe.to_csv(
        OUTPUT_DIR
        / "skill_evidence_long.csv",
        index=False,
        encoding="utf-8-sig",
    )

    full_frequency.to_csv(
        OUTPUT_DIR
        / "skill_frequency_full.csv",
        index=False,
        encoding="utf-8-sig",
    )

    core_frequency.to_csv(
        OUTPUT_DIR
        / "skill_frequency_core.csv",
        index=False,
        encoding="utf-8-sig",
    )

    role_review_dataframe.to_csv(
        OUTPUT_DIR / "role_review.csv",
        index=False,
        encoding="utf-8-sig",
    )

    anomaly_dataframe.to_csv(
        OUTPUT_DIR / "anomaly_review.csv",
        index=False,
        encoding="utf-8-sig",
    )

    if not comparison_dataframe.empty:
        comparison_dataframe.to_csv(
            OUTPUT_DIR
            / "frequency_comparison_v1_v11.csv",
            index=False,
            encoding="utf-8-sig",
        )

    # 汇总Excel
    with pd.ExcelWriter(
        OUTPUT_DIR / "skill_audit.xlsx",
        engine="openpyxl",
    ) as writer:
        jobs_dataframe.to_excel(
            writer,
            sheet_name="Jobs_Audited",
            index=False,
        )

        full_frequency.to_excel(
            writer,
            sheet_name="Skill_Frequency_Full",
            index=False,
        )

        core_frequency.to_excel(
            writer,
            sheet_name="Skill_Frequency_Core",
            index=False,
        )

        pd.concat(
            [
                full_groups,
                core_groups,
            ],
            ignore_index=True,
        ).to_excel(
            writer,
            sheet_name="Skill_Groups",
            index=False,
        )

        evidence_dataframe.to_excel(
            writer,
            sheet_name="Skill_Evidence",
            index=False,
        )

        role_review_dataframe.to_excel(
            writer,
            sheet_name="Role_Review",
            index=False,
        )

        anomaly_dataframe.to_excel(
            writer,
            sheet_name="Anomaly_Review",
            index=False,
        )

        if not comparison_dataframe.empty:
            comparison_dataframe.to_excel(
                writer,
                sheet_name="V1_V11_Comparison",
                index=False,
            )

        for worksheet in writer.book.worksheets:
            format_sheet(
                worksheet,
                {
                    "job_title": 34,
                    "company_full_name": 28,
                    "skills_v11": 60,
                    "evidence_text": 90,
                    "source_url": 45,
                    "anomaly_flags": 30,
                    "role_secondary_tags": 35,
                },
            )

    core_count = sum(
        bool(record.get("core_sample"))
        for record in audited_records
    )

    editor_trace_count = sum(
        "疑似编辑模板残留"
        in record.get(
            "anomaly_flags",
            [],
        )
        for record in audited_records
    )

    c_language_jobs = [
        record["job_title"]
        for record in audited_records
        if "C语言"
        in record.get(
            "skills_v11",
            [],
        )
    ]

    zero_skill_jobs = [
        record["job_title"]
        for record in audited_records
        if record.get(
            "skill_count_v11"
        ) == 0
    ]

    report_lines = [
        "# 岗位技能证据审计报告 v1.1",
        "",
        "## 一、审计概览",
        "",
        f"- 全样本岗位：**{len(audited_records)}**",
        f"- 核心样本岗位：**{core_count}**",
        (
            "- 因疑似编辑模板残留而排除出核心样本："
            f"**{editor_trace_count}**"
        ),
        (
            "- 需要岗位类别人工复核："
            f"**{len(role_review_dataframe)}**"
        ),
        (
            "- 低信息密度或存在异常标记："
            f"**{len(anomaly_dataframe)}**"
        ),
        (
            "- 明确识别为C语言的岗位："
            f"**{len(c_language_jobs)}**"
        ),
        (
            "- 未识别到明确技能的岗位："
            f"**{len(zero_skill_jobs)}**"
        ),
        "",
        "## 二、核心样本最高频技能",
        "",
        markdown_table(
            core_frequency,
            [
                "排名",
                "技能",
                "技能组",
                "出现岗位数",
                "岗位覆盖率百分比",
                "必备岗位数",
                "优先岗位数",
                "职责岗位数",
            ],
            20,
        ),
        "",
        "## 三、技能组覆盖情况",
        "",
        markdown_table(
            core_groups,
            [
                "技能组",
                "覆盖岗位数",
                "岗位覆盖率百分比",
            ],
            10,
        ),
        "",
        "## 四、异常样本",
        "",
        markdown_table(
            anomaly_dataframe,
            [
                "job_title",
                "jd_char_count",
                "jd_information_density",
                "anomaly_flags",
                "core_sample",
            ],
            20,
        ),
        "",
        "## 五、岗位类别复核",
        "",
        markdown_table(
            role_review_dataframe,
            [
                "job_title",
                "role_category_v1",
                "role_category_v11",
                "role_secondary_tags",
                "role_confidence",
            ],
            20,
        ),
        "",
        "## 六、审计说明",
        "",
        (
            "- 每个技能均保存命中原句，"
            "见 `skill_evidence_long.csv`。"
        ),
        (
            "- “必备、优先、职责、普通提及”"
            "依据句子中的提示词自动判断，仍属于规则推断。"
        ),
        (
            "- 核心样本只排除疑似保留编辑模板痕迹的岗位，"
            "不会删除原始记录。"
        ),
        (
            "- 该版本修复了地址中的“C座”"
            "被误判为“C语言”的问题。"
        ),
    ]

    (
        OUTPUT_DIR
        / "skill_audit_report.md"
    ).write_text(
        "\n".join(report_lines),
        encoding="utf-8",
    )

    print("=" * 72)
    print("技能证据审计 v1.1 完成")
    print("=" * 72)

    print(
        f"全样本岗位：{len(audited_records)}"
    )

    print(f"核心样本岗位：{core_count}")

    print(
        f"技能证据行数："
        f"{len(evidence_dataframe)}"
    )

    print(
        f"岗位类别待复核："
        f"{len(role_review_dataframe)}"
    )

    print(
        f"异常或低信息岗位："
        f"{len(anomaly_dataframe)}"
    )

    print(
        f"C语言命中岗位："
        f"{len(c_language_jobs)}"
    )

    print(
        f"零技能岗位："
        f"{len(zero_skill_jobs)}"
    )

    print(
        f"输出目录：{OUTPUT_DIR.resolve()}"
    )

    print()
    print("核心样本最高频技能前10名：")

    for _, row in core_frequency.head(
        10
    ).iterrows():
        print(
            f"- {row['技能']}"
            f"（{row['技能组']}）："
            f"{row['出现岗位数']} 个岗位，"
            f"{row['岗位覆盖率百分比']}"
        )


if __name__ == "__main__":
    main()
