import json
import re
from collections import Counter
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


INPUT_FILE = Path("output/boss_cleaned/jobs_cleaned.jsonl")
OUTPUT_DIR = Path("output/analysis_v1")

OUTPUT_ENRICHED_JSONL = OUTPUT_DIR / "jobs_enriched.jsonl"
OUTPUT_SUMMARY_CSV = OUTPUT_DIR / "job_summary.csv"
OUTPUT_SKILL_FREQUENCY = OUTPUT_DIR / "skill_frequency.csv"
OUTPUT_SKILL_MATRIX = OUTPUT_DIR / "job_skill_matrix.csv"
OUTPUT_DISTRIBUTIONS = OUTPUT_DIR / "sample_distributions.xlsx"
OUTPUT_EXCEL = OUTPUT_DIR / "sample_summary.xlsx"
OUTPUT_REPORT = OUTPUT_DIR / "sample_report.md"

ANALYSIS_VERSION = "1.0"


# ------------------------------------------------------------
# 技能词典
#
# 这不是从网上抄一份超大词典，而是围绕当前数据分析、AI、
# 大模型、算法和数据工程岗位样本建立的第一版可审计词典。
# ------------------------------------------------------------

SKILL_PATTERNS = {
    # 编程语言
    "Python": [
        r"(?<![a-z0-9])python(?![a-z0-9])",
    ],
    "SQL": [
        r"(?<![a-z0-9])sql(?![a-z0-9])",
    ],
    "C++": [
        r"c\+\+",
    ],
    "C语言": [
        r"(?<![a-z0-9+])c语言(?![a-z0-9])",
        r"(?<![a-z0-9+])c(?![a-z0-9+#])",
    ],
    "Java": [
        r"(?<![a-z0-9])java(?![a-z0-9])",
    ],
    "R语言": [
        r"(?<![a-z0-9])r语言(?![a-z0-9])",
        r"(?<![a-z0-9])r(?![a-z0-9])",
    ],

    # Python与数据分析生态
    "Pandas": [
        r"(?<![a-z0-9])pandas(?![a-z0-9])",
    ],
    "NumPy": [
        r"(?<![a-z0-9])numpy(?![a-z0-9])",
    ],
    "Scikit-learn": [
        r"scikit[- ]?learn",
        r"(?<![a-z0-9])sklearn(?![a-z0-9])",
    ],
    "Matplotlib": [
        r"(?<![a-z0-9])matplotlib(?![a-z0-9])",
    ],

    # 深度学习
    "PyTorch": [
        r"(?<![a-z0-9])pytorch(?![a-z0-9])",
        r"(?<![a-z0-9])torch(?![a-z0-9])",
    ],
    "TensorFlow": [
        r"tensorflow",
    ],
    "深度学习": [
        r"深度学习",
        r"deep learning",
    ],
    "机器学习": [
        r"机器学习",
        r"machine learning",
    ],
    "模型训练": [
        r"模型训练",
        r"训练模型",
    ],
    "模型部署": [
        r"模型部署",
        r"部署模型",
        r"推理部署",
    ],
    "模型评测": [
        r"模型评测",
        r"效果评测",
        r"评测脚本",
        r"模型测试",
    ],
    "时间序列": [
        r"时间序列",
        r"时序数据",
        r"高频数据",
    ],
    "量化分析": [
        r"量化模型",
        r"量化分析",
        r"因子挖掘",
        r"市场微观结构",
    ],

    # 大模型与Agent
    "大模型/LLM": [
        r"大模型",
        r"语言模型",
        r"(?<![a-z0-9])llm(?![a-z0-9])",
    ],
    "AI Agent": [
        r"ai\s*agent",
        r"智能体",
        r"agent",
    ],
    "RAG": [
        r"(?<![a-z0-9])rag(?![a-z0-9])",
        r"检索增强生成",
    ],
    "Prompt": [
        r"(?<![a-z0-9])prompt(?![a-z0-9])",
        r"提示词",
    ],
    "Fine-tuning": [
        r"fine[- ]?tune",
        r"fine[- ]?tuning",
        r"微调",
    ],
    "向量数据库": [
        r"vector database",
        r"向量数据库",
        r"向量检索",
    ],
    "知识库": [
        r"知识库",
        r"知识语料",
    ],
    "工具调用": [
        r"工具调用",
        r"tool\s*calling",
        r"skill\s*/?\s*tool",
    ],
    "工作流编排": [
        r"工作流",
        r"agent编排",
        r"流程编排",
    ],
    "多轮对话": [
        r"多轮对话",
    ],
    "意图识别": [
        r"意图识别",
        r"intent",
    ],

    # 数据分析与统计
    "Excel": [
        r"(?<![a-z0-9])excel(?![a-z0-9])",
    ],
    "数据分析": [
        r"数据分析",
        r"业务分析",
        r"分析报告",
    ],
    "统计分析": [
        r"统计分析",
        r"统计学",
    ],
    "A/B测试": [
        r"a\s*/\s*b\s*测试",
        r"ab测试",
    ],
    "假设检验": [
        r"假设检验",
    ],
    "数据可视化": [
        r"数据可视化",
        r"可视化图表",
        r"可视化分析",
    ],
    "BI": [
        r"(?<![a-z0-9])bi(?![a-z0-9])",
        r"商业智能",
        r"数据看板",
    ],
    "Tableau": [
        r"tableau",
    ],
    "FineBI": [
        r"finebi",
    ],
    "Power BI": [
        r"power\s*bi",
        r"powerbi",
    ],
    "业务指标": [
        r"业务指标",
        r"指标体系",
    ],
    "异常分析": [
        r"异常分析",
        r"异常检测",
        r"异常排查",
        r"下钻分析",
    ],

    # 数据处理
    "数据清洗": [
        r"数据清洗",
        r"清洗数据",
        r"清洗与过滤",
    ],
    "数据预处理": [
        r"数据预处理",
        r"预处理数据",
    ],
    "数据采集": [
        r"数据采集",
        r"数据收集",
        r"采集数据",
        r"抓取数据",
    ],
    "数据标注": [
        r"数据标注",
        r"标注数据",
        r"数据标签",
    ],
    "数据治理": [
        r"数据治理",
        r"数据标准化",
        r"数据质量",
        r"主数据管理",
    ],
    "数据建模": [
        r"数据建模",
        r"数据模型",
    ],
    "数据仓库": [
        r"数据仓库",
        r"数仓",
    ],
    "ETL": [
        r"(?<![a-z0-9])etl(?![a-z0-9])",
    ],
    "埋点分析": [
        r"埋点",
        r"神策",
        r"数数平台",
    ],

    # 工程技术
    "爬虫": [
        r"爬虫",
        r"网页抓取",
        r"网络抓取",
    ],
    "正则表达式": [
        r"正则",
        r"regular expression",
    ],
    "API/接口": [
        r"接口设计",
        r"api接口",
        r"(?<![a-z0-9])api(?![a-z0-9])",
    ],
    "自动化脚本": [
        r"自动化脚本",
        r"脚本开发",
    ],
    "Linux": [
        r"(?<![a-z0-9])linux(?![a-z0-9])",
    ],
    "Docker": [
        r"(?<![a-z0-9])docker(?![a-z0-9])",
    ],
    "Git": [
        r"(?<![a-z0-9])git(?![a-z0-9])",
    ],
    "Spark": [
        r"(?<![a-z0-9])spark(?![a-z0-9])",
    ],
    "Hadoop": [
        r"(?<![a-z0-9])hadoop(?![a-z0-9])",
    ],
    "Hive": [
        r"(?<![a-z0-9])hive(?![a-z0-9])",
    ],

    # 产品与协作
    "需求分析": [
        r"需求分析",
        r"需求拆解",
        r"业务需求",
    ],
    "产品设计": [
        r"产品设计",
        r"产品定义",
        r"方案设计",
    ],
    "项目管理": [
        r"项目管理",
        r"项目推进",
        r"项目落地",
    ],
    "跨团队协作": [
        r"跨团队",
        r"多团队协作",
        r"团队协作",
        r"沟通协调",
    ],
    "报告撰写": [
        r"报告撰写",
        r"撰写报告",
        r"分析报告",
        r"工作汇报",
    ],
    "英语": [
        r"英语",
        r"英文",
    ],
}


ROLE_RULES = [
    (
        "AI产品与项目",
        [
            r"产品经理",
            r"ai项目",
            r"项目实习生",
        ],
    ),
    (
        "算法与深度学习",
        [
            r"深度学习",
            r"算法实习生",
            r"算法工程师",
            r"ai算法",
        ],
    ),
    (
        "AI与大模型开发",
        [
            r"agent.*开发",
            r"开发.*agent",
            r"大模型开发",
            r"python开发",
            r"ai工程师助理",
        ],
    ),
    (
        "数据处理与标注",
        [
            r"大模型数据",
            r"数据处理",
            r"数据标注",
            r"人工智能数据分析",
            r"流程优化",
        ],
    ),
    (
        "数据采集与测试",
        [
            r"数据采集",
            r"测试数据",
        ],
    ),
    (
        "数据工程与平台",
        [
            r"数据方向",
            r"大数据",
            r"数据开发",
            r"数据仓库",
        ],
    ),
    (
        "数据分析与BI",
        [
            r"数据分析",
            r"分析师",
            r"数据应用",
            r"ai预测数据分析",
        ],
    ),
]


def clean_text(value) -> str:
    if value is None:
        return ""

    return re.sub(
        r"\s+",
        " ",
        str(value).replace("\xa0", " "),
    ).strip()


def load_jsonl(path: Path) -> list[dict]:
    if not path.exists():
        raise FileNotFoundError(
            f"找不到输入文件：{path.resolve()}"
        )

    records = []

    with path.open("r", encoding="utf-8") as file:
        for line_number, line in enumerate(file, start=1):
            line = line.strip()

            if not line:
                continue

            try:
                record = json.loads(line)
            except json.JSONDecodeError as exc:
                raise RuntimeError(
                    f"第 {line_number} 行JSON解析失败：{exc}"
                ) from exc

            records.append(record)

    if not records:
        raise RuntimeError("输入文件中没有岗位记录。")

    return records


def classify_role(title: str, description: str) -> str:
    title_text = clean_text(title).lower()
    combined = (
        title_text
        + " "
        + clean_text(description).lower()
    )

    # 优先按岗位标题判断
    for role_name, patterns in ROLE_RULES:
        for pattern in patterns:
            if re.search(pattern, title_text, re.IGNORECASE):
                return role_name

    # 标题不明确时再参考正文
    if re.search(
        r"数据仓库|数仓|数据平台|后端开发|运维开发",
        combined,
        re.IGNORECASE,
    ):
        return "数据工程与平台"

    if re.search(
        r"数据分析|业务分析|报表|看板|可视化",
        combined,
        re.IGNORECASE,
    ):
        return "数据分析与BI"

    if re.search(
        r"大模型|agent|rag|prompt",
        combined,
        re.IGNORECASE,
    ):
        return "AI与大模型开发"

    return "其他/综合"


def extract_skills(title: str, description: str) -> list[str]:
    combined = (
        clean_text(title)
        + "\n"
        + clean_text(description)
    ).lower()

    matched = []

    for skill_name, patterns in SKILL_PATTERNS.items():
        found = any(
            re.search(
                pattern,
                combined,
                re.IGNORECASE,
            )
            for pattern in patterns
        )

        if found:
            matched.append(skill_name)

    return matched


def parse_salary(salary: str) -> dict:
    salary = clean_text(salary)

    result = {
        "salary_type": "其他",
        "salary_low": None,
        "salary_high": None,
        "salary_midpoint": None,
        "salary_unit": "",
        "salary_months_per_year": None,
    }

    daily_match = re.search(
        r"(\d+(?:\.\d+)?)\s*-\s*"
        r"(\d+(?:\.\d+)?)\s*元/天",
        salary,
    )

    if daily_match:
        low = float(daily_match.group(1))
        high = float(daily_match.group(2))

        result.update(
            {
                "salary_type": "日薪",
                "salary_low": low,
                "salary_high": high,
                "salary_midpoint": round(
                    (low + high) / 2,
                    2,
                ),
                "salary_unit": "元/天",
            }
        )

        return result

    monthly_match = re.search(
        r"(\d+(?:\.\d+)?)\s*-\s*"
        r"(\d+(?:\.\d+)?)\s*[kK]"
        r"(?:\s*·\s*(\d+)\s*薪)?",
        salary,
    )

    if monthly_match:
        low = float(monthly_match.group(1))
        high = float(monthly_match.group(2))
        months = (
            int(monthly_match.group(3))
            if monthly_match.group(3)
            else 12
        )

        result.update(
            {
                "salary_type": "月薪",
                "salary_low": low,
                "salary_high": high,
                "salary_midpoint": round(
                    (low + high) / 2,
                    2,
                ),
                "salary_unit": "K/月",
                "salary_months_per_year": months,
            }
        )

        return result

    return result


def enrich_record(record: dict) -> dict:
    result = dict(record)

    title = clean_text(record.get("job_title"))
    description = clean_text(
        record.get("job_description")
    )

    skills = extract_skills(title, description)
    salary_info = parse_salary(
        record.get("salary", "")
    )

    result.update(salary_info)

    result["role_category"] = classify_role(
        title,
        description,
    )

    result["skills"] = skills
    result["skill_count"] = len(skills)
    result["skills_text"] = "；".join(skills)
    result["analysis_version"] = ANALYSIS_VERSION

    return result


def counter_dataframe(
    values,
    column_name: str,
) -> pd.DataFrame:
    cleaned_values = [
        clean_text(value) or "未披露"
        for value in values
    ]

    counter = Counter(cleaned_values)
    total = len(cleaned_values)

    rows = []

    for value, count in counter.most_common():
        rows.append(
            {
                column_name: value,
                "岗位数": count,
                "占比": round(count / total, 4),
                "占比百分比": f"{count / total:.1%}",
            }
        )

    return pd.DataFrame(rows)


def build_skill_frequency(
    records: list[dict],
) -> pd.DataFrame:
    counter = Counter()

    for record in records:
        counter.update(record["skills"])

    total_jobs = len(records)

    rows = []

    for rank, (skill, count) in enumerate(
        counter.most_common(),
        start=1,
    ):
        rows.append(
            {
                "排名": rank,
                "技能": skill,
                "出现岗位数": count,
                "岗位覆盖率": round(
                    count / total_jobs,
                    4,
                ),
                "岗位覆盖率百分比": (
                    f"{count / total_jobs:.1%}"
                ),
            }
        )

    return pd.DataFrame(rows)


def build_skill_matrix(
    records: list[dict],
    skill_frequency: pd.DataFrame,
) -> pd.DataFrame:
    skills = (
        skill_frequency["技能"].tolist()
        if not skill_frequency.empty
        else []
    )

    rows = []

    for record in records:
        row = {
            "job_id": record.get("job_id", ""),
            "job_title": record.get(
                "job_title",
                "",
            ),
            "company_full_name": record.get(
                "company_full_name",
                "",
            ),
            "role_category": record.get(
                "role_category",
                "",
            ),
        }

        record_skills = set(record["skills"])

        for skill in skills:
            row[skill] = (
                1 if skill in record_skills else 0
            )

        rows.append(row)

    return pd.DataFrame(rows)


def build_job_summary(
    records: list[dict],
) -> pd.DataFrame:
    columns = [
        "job_id",
        "job_title",
        "role_category",
        "company_full_name",
        "company_short_name",
        "city",
        "salary",
        "salary_type",
        "salary_low",
        "salary_high",
        "salary_midpoint",
        "salary_unit",
        "salary_months_per_year",
        "employment_type",
        "internship_days_per_week",
        "internship_duration",
        "experience",
        "education",
        "financing_stage",
        "company_size",
        "industry",
        "skills_text",
        "skill_count",
        "recruiter_name",
        "recruiter_title",
        "source_url",
    ]

    rows = []

    for record in records:
        rows.append(
            {
                column: record.get(column, "")
                for column in columns
            }
        )

    return pd.DataFrame(rows)


def write_jsonl(
    records: list[dict],
    output_path: Path,
) -> None:
    with output_path.open(
        "w",
        encoding="utf-8",
    ) as file:
        for record in records:
            file.write(
                json.dumps(
                    record,
                    ensure_ascii=False,
                )
                + "\n"
            )


def format_sheet(
    worksheet,
    width_overrides: dict | None = None,
) -> None:
    width_overrides = width_overrides or {}

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="0F766E",
    )

    header_font = Font(
        color="FFFFFF",
        bold=True,
    )

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.row_dimensions[1].height = 28

    for column_index, cell in enumerate(
        worksheet[1],
        start=1,
    ):
        column_name = str(cell.value)

        width = width_overrides.get(
            column_name,
            16,
        )

        worksheet.column_dimensions[
            get_column_letter(column_index)
        ].width = width

    for row in worksheet.iter_rows(
        min_row=2,
        max_row=worksheet.max_row,
    ):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )


def dataframe_to_markdown_table(
    dataframe: pd.DataFrame,
    max_rows: int = 10,
) -> str:
    if dataframe.empty:
        return "暂无数据。"

    frame = dataframe.head(max_rows).copy()

    headers = [
        str(column)
        for column in frame.columns
    ]

    lines = [
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join(
            ["---"] * len(headers)
        ) + " |",
    ]

    for _, row in frame.iterrows():
        values = []

        for value in row.tolist():
            text = clean_text(value).replace(
                "|",
                "\\|",
            )

            values.append(text)

        lines.append(
            "| " + " | ".join(values) + " |"
        )

    return "\n".join(lines)


def write_report(
    records: list[dict],
    distributions: dict[str, pd.DataFrame],
    skill_frequency: pd.DataFrame,
) -> None:
    total = len(records)

    daily_records = [
        record
        for record in records
        if record["salary_type"] == "日薪"
    ]

    monthly_records = [
        record
        for record in records
        if record["salary_type"] == "月薪"
    ]

    daily_midpoints = [
        record["salary_midpoint"]
        for record in daily_records
        if record["salary_midpoint"] is not None
    ]

    monthly_midpoints = [
        record["salary_midpoint"]
        for record in monthly_records
        if record["salary_midpoint"] is not None
    ]

    report_lines = [
        "# 个人收藏岗位画像报告",
        "",
        "## 一、样本说明",
        "",
        (
            f"本报告基于用户主动收藏并采集的 "
            f"**{total} 条 BOSS 职位信息**。"
        ),
        "",
        (
            "该样本反映的是用户当前关注方向和平台推荐结果，"
            "属于个人定向样本，不能直接推断整个招聘市场的"
            "岗位数量、薪资水平或技能需求。"
        ),
        "",
        "## 二、样本概览",
        "",
        (
            f"- 岗位总数：**{total}**"
        ),
        (
            f"- 日薪岗位：**{len(daily_records)}**"
        ),
        (
            f"- 月薪岗位：**{len(monthly_records)}**"
        ),
    ]

    if daily_midpoints:
        report_lines.append(
            "- 日薪区间中点的样本中位数："
            f"**{pd.Series(daily_midpoints).median():.0f} 元/天**"
        )

    if monthly_midpoints:
        report_lines.append(
            "- 月薪区间中点的样本中位数："
            f"**{pd.Series(monthly_midpoints).median():.1f}K/月**"
        )

    report_lines.extend(
        [
            "",
            "## 三、岗位类别",
            "",
            dataframe_to_markdown_table(
                distributions["岗位类别"][
                    [
                        "岗位类别",
                        "岗位数",
                        "占比百分比",
                    ]
                ],
                max_rows=20,
            ),
            "",
            "## 四、招聘类型",
            "",
            dataframe_to_markdown_table(
                distributions["招聘类型"][
                    [
                        "招聘类型",
                        "岗位数",
                        "占比百分比",
                    ]
                ],
                max_rows=20,
            ),
            "",
            "## 五、城市分布",
            "",
            dataframe_to_markdown_table(
                distributions["城市"][
                    [
                        "城市",
                        "岗位数",
                        "占比百分比",
                    ]
                ],
                max_rows=20,
            ),
            "",
            "## 六、学历要求",
            "",
            dataframe_to_markdown_table(
                distributions["学历"][
                    [
                        "学历",
                        "岗位数",
                        "占比百分比",
                    ]
                ],
                max_rows=20,
            ),
            "",
            "## 七、实习出勤要求",
            "",
            dataframe_to_markdown_table(
                distributions["实习出勤"][
                    [
                        "实习出勤",
                        "岗位数",
                        "占比百分比",
                    ]
                ],
                max_rows=20,
            ),
            "",
            "## 八、实习周期",
            "",
            dataframe_to_markdown_table(
                distributions["实习周期"][
                    [
                        "实习周期",
                        "岗位数",
                        "占比百分比",
                    ]
                ],
                max_rows=20,
            ),
            "",
            "## 九、最高频技能",
            "",
            dataframe_to_markdown_table(
                skill_frequency[
                    [
                        "排名",
                        "技能",
                        "出现岗位数",
                        "岗位覆盖率百分比",
                    ]
                ],
                max_rows=20,
            ),
            "",
            "## 十、使用说明",
            "",
            (
                "技能频率来自职位标题和职位描述中的明确文本匹配，"
                "表示招聘信息是否提到某项技能，不表示该技能一定是"
                "硬性要求，也不区分“必须”“优先”或“了解即可”。"
            ),
            "",
            (
                "下一阶段应人工抽查技能提取结果，再加入个人能力画像，"
                "构建可解释的岗位匹配与技能缺口分析。"
            ),
        ]
    )

    OUTPUT_REPORT.write_text(
        "\n".join(report_lines),
        encoding="utf-8",
    )


def main() -> None:
    OUTPUT_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    records = load_jsonl(INPUT_FILE)

    enriched_records = [
        enrich_record(record)
        for record in records
    ]

    skill_frequency = build_skill_frequency(
        enriched_records
    )

    skill_matrix = build_skill_matrix(
        enriched_records,
        skill_frequency,
    )

    job_summary = build_job_summary(
        enriched_records
    )

    distributions = {
        "岗位类别": counter_dataframe(
            [
                record["role_category"]
                for record in enriched_records
            ],
            "岗位类别",
        ),
        "招聘类型": counter_dataframe(
            [
                record.get("employment_type", "")
                for record in enriched_records
            ],
            "招聘类型",
        ),
        "城市": counter_dataframe(
            [
                record.get("city", "")
                for record in enriched_records
            ],
            "城市",
        ),
        "学历": counter_dataframe(
            [
                record.get("education", "")
                for record in enriched_records
            ],
            "学历",
        ),
        "实习出勤": counter_dataframe(
            [
                record.get(
                    "internship_days_per_week",
                    "",
                )
                for record in enriched_records
                if record.get(
                    "employment_type",
                    "",
                )
                in {"实习", "实习/可转正"}
            ],
            "实习出勤",
        ),
        "实习周期": counter_dataframe(
            [
                record.get(
                    "internship_duration",
                    "",
                )
                for record in enriched_records
                if record.get(
                    "employment_type",
                    "",
                )
                in {"实习", "实习/可转正"}
            ],
            "实习周期",
        ),
        "公司规模": counter_dataframe(
            [
                record.get("company_size", "")
                for record in enriched_records
            ],
            "公司规模",
        ),
        "行业": counter_dataframe(
            [
                record.get("industry", "")
                for record in enriched_records
            ],
            "行业",
        ),
        "融资阶段": counter_dataframe(
            [
                record.get("financing_stage", "")
                for record in enriched_records
            ],
            "融资阶段",
        ),
        "薪资类型": counter_dataframe(
            [
                record.get("salary_type", "")
                for record in enriched_records
            ],
            "薪资类型",
        ),
    }

    write_jsonl(
        enriched_records,
        OUTPUT_ENRICHED_JSONL,
    )

    job_summary.to_csv(
        OUTPUT_SUMMARY_CSV,
        index=False,
        encoding="utf-8-sig",
    )

    skill_frequency.to_csv(
        OUTPUT_SKILL_FREQUENCY,
        index=False,
        encoding="utf-8-sig",
    )

    skill_matrix.to_csv(
        OUTPUT_SKILL_MATRIX,
        index=False,
        encoding="utf-8-sig",
    )

    with pd.ExcelWriter(
        OUTPUT_DISTRIBUTIONS,
        engine="openpyxl",
    ) as writer:
        for sheet_name, dataframe in distributions.items():
            dataframe.to_excel(
                writer,
                sheet_name=sheet_name[:31],
                index=False,
            )

        for worksheet in writer.book.worksheets:
            format_sheet(
                worksheet,
                {
                    worksheet.cell(
                        row=1,
                        column=1,
                    ).value: 24,
                    "岗位数": 12,
                    "占比": 14,
                    "占比百分比": 16,
                },
            )

    with pd.ExcelWriter(
        OUTPUT_EXCEL,
        engine="openpyxl",
    ) as writer:
        job_summary.to_excel(
            writer,
            sheet_name="Job_Summary",
            index=False,
        )

        skill_frequency.to_excel(
            writer,
            sheet_name="Skill_Frequency",
            index=False,
        )

        skill_matrix.to_excel(
            writer,
            sheet_name="Job_Skill_Matrix",
            index=False,
        )

        for sheet_name, dataframe in distributions.items():
            dataframe.to_excel(
                writer,
                sheet_name=sheet_name[:31],
                index=False,
            )

        workbook = writer.book

        format_sheet(
            workbook["Job_Summary"],
            {
                "job_id": 26,
                "job_title": 34,
                "role_category": 20,
                "company_full_name": 30,
                "company_short_name": 22,
                "city": 10,
                "salary": 16,
                "salary_type": 12,
                "salary_low": 12,
                "salary_high": 12,
                "salary_midpoint": 14,
                "salary_unit": 12,
                "salary_months_per_year": 18,
                "employment_type": 16,
                "internship_days_per_week": 18,
                "internship_duration": 16,
                "experience": 16,
                "education": 14,
                "financing_stage": 16,
                "company_size": 16,
                "industry": 20,
                "skills_text": 60,
                "skill_count": 12,
                "recruiter_name": 14,
                "recruiter_title": 20,
                "source_url": 45,
            },
        )

        format_sheet(
            workbook["Skill_Frequency"],
            {
                "排名": 10,
                "技能": 24,
                "出现岗位数": 14,
                "岗位覆盖率": 16,
                "岗位覆盖率百分比": 20,
            },
        )

        skill_matrix_widths = {
            "job_id": 26,
            "job_title": 34,
            "company_full_name": 28,
            "role_category": 20,
        }

        for skill in skill_frequency["技能"].tolist():
            skill_matrix_widths[skill] = 14

        format_sheet(
            workbook["Job_Skill_Matrix"],
            skill_matrix_widths,
        )

        for sheet_name in distributions:
            format_sheet(
                workbook[sheet_name[:31]],
            )

    write_report(
        enriched_records,
        distributions,
        skill_frequency,
    )

    print("=" * 72)
    print("第一版岗位画像分析完成")
    print("=" * 72)
    print(f"输入岗位数：{len(records)}")
    print(f"识别技能数：{len(skill_frequency)}")
    print(
        "平均每个岗位识别技能数："
        f"{pd.Series([r['skill_count'] for r in enriched_records]).mean():.2f}"
    )
    print()
    print(f"岗位汇总：{OUTPUT_SUMMARY_CSV.resolve()}")
    print(f"技能频率：{OUTPUT_SKILL_FREQUENCY.resolve()}")
    print(f"技能矩阵：{OUTPUT_SKILL_MATRIX.resolve()}")
    print(f"分析Excel：{OUTPUT_EXCEL.resolve()}")
    print(f"分布Excel：{OUTPUT_DISTRIBUTIONS.resolve()}")
    print(f"中文报告：{OUTPUT_REPORT.resolve()}")

    print()
    print("最高频技能前10名：")

    for _, row in skill_frequency.head(10).iterrows():
        print(
            f"- {row['技能']}："
            f"{row['出现岗位数']} 个岗位，"
            f"{row['岗位覆盖率百分比']}"
        )


if __name__ == "__main__":
    main()
