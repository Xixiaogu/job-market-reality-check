import asyncio
import json
import re
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from playwright.async_api import async_playwright


EXCEL_FILE = Path("BOSS链接.xlsx")
CDP_URL = "http://127.0.0.1:9222"
OUTPUT_DIR = Path("output") / "single_job_structured"

URL_PATTERN = re.compile(r"https?://[^\s<>\"']+")
JOB_ID_PATTERN = re.compile(r"/weijd-job/([^?/#]+)")


def clean_inline(text: str | None) -> str:
    """清理单行字段中的多余空格和换行。"""
    if not text:
        return ""

    return re.sub(r"\s+", " ", text).strip()


def clean_multiline(text: str | None) -> str:
    """保留多行结构，同时清理每行多余空格。"""
    if not text:
        return ""

    lines = []

    for line in text.replace("\xa0", " ").splitlines():
        line = re.sub(r"[ \t]+", " ", line).strip()

        if line:
            lines.append(line)

    return "\n".join(lines)


def split_middle_dot(text: str) -> list[str]:
    """按照中文间隔点拆分复合字段。"""
    if not text:
        return []

    return [
        part.strip()
        for part in re.split(r"\s*[·•]\s*", text)
        if part.strip()
    ]


def read_first_job_url() -> str:
    """从Excel中读取第一条BOSS岗位详情链接。"""
    if not EXCEL_FILE.exists():
        raise FileNotFoundError(
            f"找不到Excel文件：{EXCEL_FILE.resolve()}"
        )

    workbook = load_workbook(
        EXCEL_FILE,
        read_only=True,
        data_only=True,
    )

    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str):
                    continue

                for url in URL_PATTERN.findall(cell.value):
                    if (
                        "zhipin.com" in url
                        and "/weijd/weijd-job/" in url
                    ):
                        return url

    raise RuntimeError("Excel中没有找到BOSS岗位详情链接。")


async def get_text(
    page,
    selector: str,
    *,
    use_text_content: bool = False,
    multiline: bool = False,
) -> str:
    """安全读取一个DOM元素的文字。"""
    locator = page.locator(selector).first

    if await locator.count() == 0:
        return ""

    try:
        if use_text_content:
            value = await locator.text_content(timeout=10_000)
        else:
            value = await locator.inner_text(timeout=10_000)
    except Exception:
        try:
            value = await locator.text_content(timeout=5_000)
        except Exception:
            return ""

    if multiline:
        return clean_multiline(value)

    return clean_inline(value)


async def find_or_open_job_page(browser, job_url: str):
    """优先使用已经打开的目标岗位页面，否则在当前Chrome中打开。"""
    match = JOB_ID_PATTERN.search(job_url)

    if not match:
        raise RuntimeError(f"无法从链接提取岗位ID：{job_url}")

    target_job_id = match.group(1)

    all_pages = []

    for context in browser.contexts:
        all_pages.extend(context.pages)

    for page in reversed(all_pages):
        if (
            target_job_id in page.url
            and "/weijd/weijd-job/" in page.url
        ):
            print("找到已经打开的目标岗位页面。")
            return page, target_job_id

    if not browser.contexts:
        raise RuntimeError("Chrome中没有可用的浏览器上下文。")

    context = browser.contexts[0]

    page = (
        all_pages[-1]
        if all_pages
        else await context.new_page()
    )

    print("目标岗位尚未打开，正在访问Excel第一条链接……")

    await page.goto(
        job_url,
        wait_until="domcontentloaded",
        timeout=60_000,
    )

    return page, target_job_id


async def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    job_url = read_first_job_url()

    print("=" * 70)
    print("准备提取Excel第一条岗位：")
    print(job_url)
    print("=" * 70)

    playwright = await async_playwright().start()

    try:
        print("正在连接远程调试Chrome……")

        browser = await playwright.chromium.connect_over_cdp(
            CDP_URL
        )

        page, expected_job_id = await find_or_open_job_page(
            browser,
            job_url,
        )

        await page.bring_to_front()

        await page.wait_for_selector(
            ".boss-info",
            state="attached",
            timeout=30_000,
        )

        await page.wait_for_selector(
            ".rec-position",
            state="attached",
            timeout=30_000,
        )

        await page.wait_for_timeout(1_500)

        final_url = page.url

        if "/weijd/weijd-job/" not in final_url:
            raise RuntimeError(
                "当前页面不是岗位详情页，拒绝提取：\n"
                f"{final_url}"
            )

        final_job_match = JOB_ID_PATTERN.search(final_url)
        final_job_id = (
            final_job_match.group(1)
            if final_job_match
            else expected_job_id
        )

        # --------------------------------------------------
        # 1. 直接读取红框主体
        # --------------------------------------------------

        core_text = await get_text(
            page,
            ".boss-info",
            multiline=True,
        )

        # --------------------------------------------------
        # 2. 招聘者字段
        # --------------------------------------------------

        recruiter_name = await get_text(
            page,
            ".boss-info .boss-message .name > p",
        )

        influence = await get_text(
            page,
            ".boss-info .boss-message .name > a",
        )

        recruiter_info_raw = await get_text(
            page,
            ".boss-info .boss-message .info-labels > span",
        )

        recruiter_parts = split_middle_dot(
            recruiter_info_raw
        )

        recruiter_title = (
            recruiter_parts[0]
            if len(recruiter_parts) >= 1
            else ""
        )

        recruiter_company = (
            "·".join(recruiter_parts[1:])
            if len(recruiter_parts) >= 2
            else ""
        )

        # --------------------------------------------------
        # 3. 岗位基础字段
        # 注意：网页源码里确实叫 job-cotent
        # --------------------------------------------------

        job_title = await get_text(
            page,
            ".rec-position .job-cotent h2 > p",
        )

        salary = await get_text(
            page,
            ".rec-position .job-cotent > div",
        )

        company_short_stage_raw = await get_text(
            page,
            ".rec-position .job-brandComInfo",
            multiline=True,
        )

        company_short_stage_lines = [
            clean_inline(line)
            for line in company_short_stage_raw.splitlines()
            if clean_inline(line)
        ]

        company_short_name = (
            company_short_stage_lines[0]
            if company_short_stage_lines
            else ""
        )

        financing_stage_from_title = (
            company_short_stage_lines[1]
            if len(company_short_stage_lines) >= 2
            else ""
        )

        job_basic_info_raw = await get_text(
            page,
            ".rec-position .job-cotent "
            "h2 > span:not(.job-brandComInfo)",
        )

        basic_parts = split_middle_dot(
            job_basic_info_raw
        )

        city = (
            basic_parts[0]
            if len(basic_parts) >= 1
            else ""
        )

        experience = (
            basic_parts[1]
            if len(basic_parts) >= 2
            else ""
        )

        education = (
            "·".join(basic_parts[2:])
            if len(basic_parts) >= 3
            else ""
        )

        # --------------------------------------------------
        # 4. 职位详情
        # 完整正文已在DOM中，不点击“查看全部”
        # --------------------------------------------------

        job_description = await get_text(
            page,
            ".rec-position .rec-detail .detail-text",
            use_text_content=True,
            multiline=True,
        )

        job_description = re.sub(
            r"(?:\n)?查看全部\s*$",
            "",
            job_description,
        ).strip()

        work_address = await get_text(
            page,
            ".rec-position .detail-text.address",
            multiline=True,
        )

        # --------------------------------------------------
        # 5. 公司字段
        # --------------------------------------------------

        company_full_name = await get_text(
            page,
            ".rec-position .brandComBaseInfo > span",
        )

        company_info_raw = await get_text(
            page,
            ".rec-position .brandComBaseInfo > p",
        )

        company_parts = split_middle_dot(
            company_info_raw
        )

        financing_stage = (
            company_parts[0]
            if len(company_parts) >= 1
            else financing_stage_from_title
        )

        company_size = (
            company_parts[1]
            if len(company_parts) >= 2
            else ""
        )

        industry = (
            "·".join(company_parts[2:])
            if len(company_parts) >= 3
            else ""
        )

        # --------------------------------------------------
        # 6. 汇总结构化结果
        # --------------------------------------------------

        result = {
            "job_id": final_job_id,
            "recruiter_name": recruiter_name,
            "influence": influence,
            "recruiter_title": recruiter_title,
            "recruiter_company": recruiter_company,
            "job_title": job_title,
            "salary": salary,
            "city": city,
            "experience": experience,
            "education": education,
            "company_short_name": company_short_name,
            "financing_stage": financing_stage,
            "job_description": job_description,
            "work_address": work_address,
            "company_full_name": company_full_name,
            "company_size": company_size,
            "industry": industry,
            "recruiter_info_raw": recruiter_info_raw,
            "job_basic_info_raw": job_basic_info_raw,
            "company_info_raw": company_info_raw,
            "core_text": core_text,
            "source_url": job_url,
            "final_url": final_url,
            "collected_at": datetime.now()
            .astimezone()
            .isoformat(),
            "status": "success",
        }

        # --------------------------------------------------
        # 7. 保存结果
        # --------------------------------------------------

        html = await page.content()

        (OUTPUT_DIR / "job.json").write_text(
            json.dumps(
                result,
                ensure_ascii=False,
                indent=2,
            ),
            encoding="utf-8",
        )

        (OUTPUT_DIR / "core_text.txt").write_text(
            core_text,
            encoding="utf-8",
        )

        (OUTPUT_DIR / "page.html").write_text(
            html,
            encoding="utf-8",
        )

        dataframe = pd.DataFrame([result])

        dataframe.to_excel(
            OUTPUT_DIR / "job.xlsx",
            index=False,
            engine="openpyxl",
        )

        dataframe.to_csv(
            OUTPUT_DIR / "job.csv",
            index=False,
            encoding="utf-8-sig",
        )

        await page.screenshot(
            path=str(OUTPUT_DIR / "page.png"),
            full_page=True,
        )

        # --------------------------------------------------
        # 8. 控制台预览
        # --------------------------------------------------

        print("\n" + "=" * 70)
        print("岗位主体提取成功")
        print("=" * 70)

        print(f"招聘者：{recruiter_name}")
        print(f"招聘者职位：{recruiter_title}")
        print(f"岗位名称：{job_title}")
        print(f"薪资：{salary}")
        print(
            f"基础要求：{city} / "
            f"{experience} / {education}"
        )
        print(f"公司：{company_full_name}")
        print(f"公司信息：{company_info_raw}")
        print(f"岗位描述长度：{len(job_description)}")
        print(f"最终页面：{final_url}")
        print(f"输出目录：{OUTPUT_DIR.resolve()}")

        print("\n岗位描述预览：")
        print("-" * 70)
        print(job_description[:1200])
        print("-" * 70)

        # 不执行 browser.close()
        # 避免关闭用户当前正在使用的Chrome

    finally:
        await playwright.stop()


if __name__ == "__main__":
    asyncio.run(main())
